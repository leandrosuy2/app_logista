from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import authenticate, login  
from .models import Devedor, Empresa, Titulo, Acordo, Parcelamento, UserAccessLog, MensagemWhatsapp, TabelaRemuneracao, TabelaRemuneracaoLista
from django.apps import AppConfig
from core.models import UserLojista
from django.contrib.auth.hashers import check_password
from django.core.paginator import Paginator
from django.db.models.functions import TruncDate
import logging
import bcrypt
from datetime import date, datetime
from core.models import UserLojista
from django.apps import AppConfig
from django.urls import reverse
from django.contrib import messages
from .models import Empresa, Parcelamento, FollowUp
from django.db import connection
from django.http import HttpResponseNotFound, JsonResponse, HttpResponse
from dateutil.relativedelta import relativedelta
from django.utils.dateformat import format
from django.db.models import F, Q, Sum 
from django.db.models import Case, When, IntegerField, Q, F
from django.views.decorators.http import require_POST
from core.models import Acordo, TipoDocTitulo, Agendamento, FollowUp
from django.utils.timezone import make_aware, now, localtime
import re
from .utils import consultar_cnpj_via_scraping
import time
from django.contrib.auth.models import User, Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.hashers import make_password
from django.utils.translation import gettext_lazy as _
from django.views.decorators.csrf import csrf_exempt
import json
# WeasyPrint - importação opcional (requer bibliotecas GTK no Windows)
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError):
    WEASYPRINT_AVAILABLE = False
    HTML = None  # Placeholder para evitar erros

from django.template.loader import render_to_string
import tempfile
import traceback
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from functools import wraps
from django.shortcuts import redirect
from types import SimpleNamespace

def minha_view(request):
    return redirect('https://lojista.intercred.com.br')


from django.contrib.auth.decorators import login_required, permission_required
from django.db import connection
from num2words import num2words
import os
from django.conf import settings
from .forms import MensagemWhatsappForm
from django.core.exceptions import ValidationError
from decimal import Decimal
import uuid


 


import logging

logger = logging.getLogger(__name__)

# --- helpers de protocolo ---
from django.utils import timezone
from django.db import connection
import random

_PROTO_FIELD_CACHE = None

def _get_protocolo_field():
    """
    Retorna o nome da coluna de protocolo existente na tabela titulo.
    Prioridade: protocolo > codigo_protocolo > protocolo_gerado.
    """
    global _PROTO_FIELD_CACHE
    if _PROTO_FIELD_CACHE:
        return _PROTO_FIELD_CACHE

    with connection.cursor() as cur:
        cur.execute("""
            SELECT COLUMN_NAME
            FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = 'titulo'
              AND COLUMN_NAME IN ('protocolo', 'codigo_protocolo', 'protocolo_gerado')
            ORDER BY FIELD(COLUMN_NAME,'protocolo','codigo_protocolo','protocolo_gerado')
            LIMIT 1
        """)
        row = cur.fetchone()
        _PROTO_FIELD_CACHE = row[0] if row else None
    return _PROTO_FIELD_CACHE

def _gerar_protocolo(titulo_id: int) -> str:
    # Ex.: NC-20250829-000123-4821
    return f"NC-{timezone.now().strftime('%Y%m%d')}-{titulo_id:06d}-{random.randint(1000,9999)}"

def set_protocolo_if_missing(titulo_obj):
    """
    Define o protocolo no objeto Titulo se a coluna existir e estiver vazia.
    Salva apenas o campo de protocolo.
    """
    field = _get_protocolo_field()
    if not field:
        return  # não há coluna de protocolo, nada a fazer
    if getattr(titulo_obj, field, None):
        return  # já tem protocolo
    setattr(titulo_obj, field, _gerar_protocolo(titulo_obj.id))
    titulo_obj.save(update_fields=[field])


def lojista_login_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.session.get('is_authenticated'):
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def home_redirect(request):
    if request.user.is_authenticated:
        return redirect('dashboard')  # Redireciona apenas para usuários autenticados
    return redirect('login')  # Redireciona para login se não estiver autenticado

    

class CoreConfig(AppConfig):
    default_auto_field = 'django.db.models.BigAutoField'
    name = 'core'

 
def format_whatsapp_number(phone):
    """Formata o número de telefone para o padrão do WhatsApp (sem caracteres especiais e com prefixo 55)."""
    if not phone:
        return None
    # Remove caracteres não numéricos
    phone = re.sub(r'\D', '', phone)
    # Adiciona o código do Brasil (55) se o número não começar com ele
    if not phone.startswith('55'):
        phone = f'55{phone}'
    return phone



# views.py (topo)
from django.http import JsonResponse, Http404
from django.utils.timezone import now, localtime

from .models import Titulo, Agendamento, FollowUp, Devedor  # <— garante Devedor aqui

# views.py — endpoint para o modal do lojista
@lojista_login_required
def followups_devedor_json(request, devedor_id: int):
    empresa_id = request.session.get('empresa_id_sessao')
    # valida se o devedor é da empresa logada
    devedor = get_object_or_404(Devedor, id=devedor_id, empresa_id=empresa_id)

    try:
        limit = int(request.GET.get('limit', 50))
    except Exception:
        limit = 50
    limit = max(1, min(limit, 500))

    itens = (FollowUp.objects
             .filter(devedor_id=devedor.id)
             .order_by('-created_at')[:limit])

    data = [{
        'created_at': localtime(fu.created_at).strftime('%d/%m/%Y %H:%M') if fu.created_at else '',
        'texto': fu.texto or ''
    } for fu in itens]

    return JsonResponse({
        'devedor_id': devedor.id,
        'count': len(data),
        'items': data
    })

@lojista_login_required
def dashboard(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    print("Empresa ID da sessão:", empresa_id_sessao) 
    hoje = now().date()
    query = request.GET.get('query', '').strip()

    # Relatórios rápidos
    titulos_pendentes = Titulo.objects.filter(
    Q(statusBaixa=0) | Q(statusBaixa__isnull=True),
    devedor__empresa_id=empresa_id_sessao
).count()
    titulos_quitados = Titulo.objects.filter(
    statusBaixa=2,
    devedor__empresa_id=empresa_id_sessao
).count()
    titulos_negociados = Titulo.objects.filter(
    statusBaixa=3,
    devedor__empresa_id=empresa_id_sessao
).count()

    total_clientes = Devedor.objects.filter(empresa_id=empresa_id_sessao).count()


    # Consulta para "Negociados Hoje"
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT SUM(titulo.valor) AS total_negociados_hoje
            FROM titulo, devedores, core_empresa
            WHERE
            titulo.devedor_id=devedores.id and devedores.empresa_id=core_empresa.id and
            titulo.statusBaixa = 3 AND titulo.created_at LIKE CONCAT(CURDATE(), '%%')
            AND devedores.empresa_id = %s;
        """,[empresa_id_sessao])
        result = cursor.fetchone()
        negociados_hoje = Decimal(result[0]) if result[0] is not None else Decimal('0.00')
        negociados_hoje = round(negociados_hoje, 2)

    # Consulta para "Quitados Hoje"
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT SUM(COALESCE(titulo.valorRecebido, 0)) AS total_quitados_hoje
            FROM titulo, devedores, core_empresa
            WHERE titulo.data_baixa = CURDATE()
            AND titulo.devedor_id=devedores.id and devedores.empresa_id=core_empresa.id
            AND devedores.empresa_id = %s;
        """, [empresa_id_sessao])
        result = cursor.fetchone()
        quitados_hoje = Decimal(result[0]) if result[0] is not None else Decimal('0.00')
        quitados_hoje = round(quitados_hoje, 2)

    # Detalhes de "Quitados Hoje"
    quitados_hoje_detalhes = Titulo.objects.raw("""
          SELECT titulo.id, devedores.nome, COALESCE(devedores.cpf, devedores.cnpj) AS cpf_cnpj,
               core_empresa.nome_fantasia, titulo.data_baixa, titulo.valorRecebido
        FROM titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE titulo.data_baixa = %s AND devedores.empresa_id = %s
    """, [hoje, empresa_id_sessao])

    quitados_hoje_detalhes_data = [{
        'nome': q.nome,
        'cpf_cnpj': q.cpf_cnpj,
        'nome_fantasia': q.nome_fantasia,
        'data_baixa': q.data_baixa.strftime('%d/%m/%Y') if q.data_baixa else '-',
        'valorRecebido': f'R$ {float(q.valorRecebido):,.2f}' if q.valorRecebido else 'R$ 0.00'
    } for q in quitados_hoje_detalhes]

    # Detalhes de "Negociados Hoje"
    negociados_hoje_detalhes = Titulo.objects.raw("""
        SELECT titulo.id, devedores.nome, COALESCE(devedores.cpf, devedores.cnpj) AS cpf_cnpj,
               core_empresa.nome_fantasia, titulo.created_at, titulo.valor
        FROM titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE titulo.statusBaixa = 3 AND DATE(titulo.created_at) = %s
    """, [hoje])

    negociados_hoje_detalhes_data = [{
        'nome': n.nome,
        'cpf_cnpj': n.cpf_cnpj,
        'nome_fantasia': n.nome_fantasia,
        'data_negociacao': n.created_at.strftime('%d/%m/%Y %H:%M') if n.created_at else '-',
        'valor': f'R$ {float(n.valor):,.2f}' if n.valor else 'R$ 0.00'
    } for n in negociados_hoje_detalhes]

    # Contagem de negociados em atraso
    negociados_em_atraso_count = Titulo.objects.filter(
    statusBaixa=3,
    dataVencimento__lt=hoje,
    devedor__empresa_id=empresa_id_sessao
).count()


    # Parcelamentos pendentes atrasados ou vencendo hoje
    parcelamentos_atrasados = Parcelamento.objects.filter(
        Q(
            Q(status='Pendente') &
            Q(data_vencimento_parcela__lte=hoje) &
            ~Q(acordo__titulo__ultima_acao=hoje)
        )
    ).select_related(
        'acordo', 'acordo__titulo', 'acordo__titulo__devedor', 'acordo__titulo__devedor__empresa'
    ).annotate(
        qtde_prc=F('acordo__qtde_prc')
    )

    # Últimas movimentações
    ultimos_movimentos = Acordo.objects.select_related('devedor', 'titulo').order_by('-id')[:10].values(
        'id',
        'devedor__nome',
        'titulo_id',
        'entrada',
        'data_entrada',
        'contato',
    )

    # Últimos clientes cadastrados
    ultimos_clientes = Devedor.objects.filter(
    empresa_id=empresa_id_sessao
).order_by('-id')[:10].values(
    'id',
    'nome',
    'cpf',
    'cnpj',
    'created_at',
)


    # Agendamentos do dia corrente
    hoje = localtime(now()).date()  # Obtém a data local corretamente

    agendamentos_hoje = Agendamento.objects.filter(
        status='Pendente',
        empresa_id=empresa_id_sessao,
        data_retorno__date=hoje  # Filtra pela data local
    ).select_related('devedor', 'empresa').values(
        'id',
        'devedor__nome',
        'devedor__cpf',
        'devedor__cnpj',
        'empresa__nome_fantasia',
        'telefone',
        'data_retorno',
        'data_abertura',
        'assunto',
        'operador',
        'status'
    )

    # Filtro de busca para pendentes
    search_filter = ""
    params = []
    if query:
        search_filter = """
        AND (
            devedores.nome LIKE %s OR
            core_empresa.nome_fantasia LIKE %s OR
            devedores.cpf LIKE %s OR
            devedores.cnpj LIKE %s
        )
        """
        params = [f"%{query}%"] * 4

    agenda_pendentes_query = f"""
    SELECT    
        titulo.id AS id,
        devedores.nome,    
        core_empresa.nome_fantasia AS nome_fantasia_credor,
        devedores.nome_mae,
        devedores.cpf,
        devedores.cnpj,
        devedores.rg,    
        devedores.telefone1
    FROM 
        devedores, titulo, core_empresa
    WHERE
        titulo.devedor_id = devedores.id 
        AND devedores.empresa_id = core_empresa.id 
        AND devedores.empresa_id = %s
        AND (titulo.statusBaixa=0 OR titulo.statusBaixa IS NULL)
        AND (titulo.ultima_acao IS NULL OR DATE(titulo.ultima_acao) != CURDATE())
        {search_filter}
    GROUP BY 
        titulo.id,
        devedores.nome, 
        core_empresa.nome_fantasia, 
        devedores.nome_mae, 
        devedores.cpf, 
        devedores.cnpj, 
        devedores.rg,
        titulo.juros,        
        devedores.telefone1
    ORDER BY 
        titulo.id DESC
    """
    agenda_pendentes = Titulo.objects.raw(agenda_pendentes_query, [empresa_id_sessao, *params])


    # Paginação para Agenda de Pendentes
    paginator_pendentes = Paginator(agenda_pendentes, 10)
    page_number_pendentes = request.GET.get('page')
    agenda_pendentes_paginated = paginator_pendentes.get_page(page_number_pendentes)

    # Filtro de busca para negociados
    negociados_em_atraso_query = f"""
    SELECT
        MIN(titulo.id) AS id,
        core_empresa.id AS empresa_id,
        devedores.nome AS devedor_nome,
        core_empresa.nome_fantasia AS empresa_nome,
        devedores.nome_mae AS nome_mae,
        titulo.devedor_id AS devedor_id,
        MIN(titulo.dataVencimento) AS data_vencimento,
        SUM(titulo.valor) AS valor_total
    FROM 
        titulo
    JOIN 
        devedores ON titulo.devedor_id = devedores.id
    JOIN 
        core_empresa ON devedores.empresa_id = core_empresa.id
    WHERE 
        titulo.statusBaixa = 3 
        AND titulo.dataVencimento < CURRENT_DATE
        AND devedores.empresa_id = {empresa_id_sessao}
        {search_filter}
    GROUP BY 
        core_empresa.id, 
        devedores.nome, 
        devedores.nome_mae, 
        titulo.devedor_id, 
        core_empresa.nome_fantasia
        """
    negociados_em_atraso = Titulo.objects.raw(negociados_em_atraso_query, params)

    # Paginação para Negociados em Atraso
    paginator_negociados = Paginator(list(negociados_em_atraso), 10)
    page_number_negociados = request.GET.get('page_negociados')
    negociados_paginated = paginator_negociados.get_page(page_number_negociados)

    # Buscar cobranças pendentes da tabela core_cobranca
    cobrancas_pendentes = []
    total_comissao_cobrancas = Decimal('0.00')
    
    try:
        with connection.cursor() as cursor:
            # Primeiro, verificar se a tabela existe e quais colunas tem
            cursor.execute("""
                SELECT COLUMN_NAME 
                FROM information_schema.COLUMNS 
                WHERE TABLE_SCHEMA = DATABASE() 
                AND TABLE_NAME = 'core_cobranca'
            """)
            columns_info = [row[0] for row in cursor.fetchall()]
            
            if columns_info:  # Se a tabela existe
                # Verificar se existe campo comissao ou valor_comissao
                campo_comissao = 'comissao' if 'comissao' in columns_info else ('valor_comissao' if 'valor_comissao' in columns_info else None)
                
                # Verificar se existe campo pago
                tem_campo_pago = 'pago' in columns_info
                
                # Montar query com campo de comissão se existir
                campos_select = "id, data_cobranca, tipo_anexo, documento, link, created_at, updated_at, empresa_id"
                if campo_comissao:
                    campos_select += f", {campo_comissao}"
                if tem_campo_pago:
                    campos_select += ", pago"
                
                # Montar filtro WHERE
                filtro_where = "WHERE empresa_id = %s"
                if tem_campo_pago:
                    filtro_where += " AND (pago != 1 OR pago IS NULL)"
                
                # Buscar cobranças (apenas as não pagas)
                cursor.execute(f"""
                    SELECT {campos_select}
                    FROM core_cobranca
                    {filtro_where}
                    ORDER BY data_cobranca DESC, created_at DESC
                """, [empresa_id_sessao])
                
                columns = [col[0] for col in cursor.description]
                for row in cursor.fetchall():
                    cobranca = dict(zip(columns, row))
                    # Converter data_cobranca para string formatada se necessário
                    if cobranca.get('data_cobranca'):
                        if isinstance(cobranca['data_cobranca'], (date, datetime)):
                            if isinstance(cobranca['data_cobranca'], datetime):
                                data_obj = cobranca['data_cobranca'].date()
                            else:
                                data_obj = cobranca['data_cobranca']
                            cobranca['data_cobranca'] = data_obj.strftime('%d/%m/%Y')
                            cobranca['data_cobranca_raw'] = data_obj.strftime('%Y-%m-%d')  # Para cálculos
                        elif isinstance(cobranca['data_cobranca'], str):
                            # Se já for string, tentar converter e formatar
                            try:
                                if len(cobranca['data_cobranca']) == 10 and '-' in cobranca['data_cobranca']:
                                    data_obj = datetime.strptime(cobranca['data_cobranca'], '%Y-%m-%d').date()
                                    cobranca['data_cobranca'] = data_obj.strftime('%d/%m/%Y')
                                    cobranca['data_cobranca_raw'] = data_obj.strftime('%Y-%m-%d')
                            except:
                                pass
                    
                    # Buscar valor da comissão
                    # Se já existe campo de comissão na tabela, usar ele
                    if campo_comissao and cobranca.get(campo_comissao):
                        try:
                            comissao_valor = Decimal(str(cobranca[campo_comissao]))
                        except:
                            comissao_valor = Decimal('0.00')
                    else:
                        # Calcular 15% dos títulos quitados desde a data da cobrança
                        data_cobranca_raw = cobranca.get('data_cobranca_raw')
                        if data_cobranca_raw:
                            try:
                                if isinstance(data_cobranca_raw, str):
                                    data_cobranca_obj = datetime.strptime(data_cobranca_raw, '%Y-%m-%d').date()
                                else:
                                    data_cobranca_obj = data_cobranca_raw
                            except:
                                data_cobranca_obj = hoje
                        else:
                            # Tentar usar data_cobranca original
                            data_cobranca = cobranca.get('data_cobranca')
                            if data_cobranca:
                                try:
                                    if isinstance(data_cobranca, str) and '-' in data_cobranca:
                                        data_cobranca_obj = datetime.strptime(data_cobranca, '%Y-%m-%d').date()
                                    elif isinstance(data_cobranca, (date, datetime)):
                                        if isinstance(data_cobranca, datetime):
                                            data_cobranca_obj = data_cobranca.date()
                                        else:
                                            data_cobranca_obj = data_cobranca
                                    else:
                                        data_cobranca_obj = hoje
                                except:
                                    data_cobranca_obj = hoje
                            else:
                                data_cobranca_obj = hoje
                        
                        cursor.execute("""
                            SELECT SUM(COALESCE(titulo.valorRecebido, 0)) * 0.15 AS comissao
                            FROM titulo
                            INNER JOIN devedores ON titulo.devedor_id = devedores.id
                            WHERE devedores.empresa_id = %s
                            AND titulo.statusBaixa = 2
                            AND titulo.data_baixa >= %s
                        """, [empresa_id_sessao, data_cobranca_obj])
                        
                        comissao_result = cursor.fetchone()
                        comissao_valor = Decimal(str(comissao_result[0])) if comissao_result and comissao_result[0] else Decimal('0.00')
                    
                    cobranca['comissao'] = round(comissao_valor, 2)
                    total_comissao_cobrancas += comissao_valor
                    
                    cobrancas_pendentes.append(cobranca)
    except Exception as e:
        # Se a tabela não existir ou houver erro, simplesmente não mostra cobranças
        logger.error(f"Erro ao buscar cobranças: {str(e)}")
        cobrancas_pendentes = []
    
    total_comissao_cobrancas = round(total_comissao_cobrancas, 2)

    # Contexto final
    context = {
        'titulos_pendentes': titulos_pendentes,
        'titulos_quitados': titulos_quitados,
        'titulos_negociados': titulos_negociados,
        'total_clientes': total_clientes,
        'negociados_em_atraso_count': negociados_em_atraso_count,
        'parcelamentos_atrasados': parcelamentos_atrasados,
        'ultimos_movimentos': ultimos_movimentos,
        'ultimos_clientes': ultimos_clientes,
        'agendamentos_hoje': agendamentos_hoje,
        'agenda_pendentes_paginated': agenda_pendentes_paginated,
        'negociados_paginated': negociados_paginated,
        'query': query,
        'quitados_hoje': quitados_hoje,
        'negociados_hoje': negociados_hoje,
        'quitados_hoje_detalhes': quitados_hoje_detalhes_data,
        'negociados_hoje_detalhes': negociados_hoje_detalhes_data,
        'cobrancas_pendentes': cobrancas_pendentes,
        'total_comissao_cobrancas': total_comissao_cobrancas,
    }

    return render(request, 'dashboard.html', context)




@lojista_login_required

def listar_grupos(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    print("Empresa ID da sessão:", empresa_id_sessao) 
    grupos = Group.objects.all()
    return render(request, 'grupos_listar.html', {'grupos': grupos})

# Criar os grupos e permissões (executar uma vez ou em um script separado)
@lojista_login_required

def criar_grupos():
    # Criar ou obter os grupos
    admin_group, _ = Group.objects.get_or_create(name='Admin')
    lojista_group, _ = Group.objects.get_or_create(name='Lojista')
    operador_group, _ = Group.objects.get_or_create(name='Operador')

    print("Grupos criados ou já existentes:")
    print(f" - {admin_group}")
    print(f" - {lojista_group}")
    print(f" - {operador_group}")

@lojista_login_required

def editar_grupo(request, grupo_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    print("Empresa ID da sessão:", empresa_id_sessao) 
    grupo = Group.objects.get(id=grupo_id)
    todas_permissoes = Permission.objects.all()  # Todas as permissões disponíveis

    # Associa permissões ao grupo (verifica se estão atribuídas)
    permissoes_detalhadas = [
        {
            'id': permissao.id,
            'codename': permissao.codename,
            'traduzido': permissao.name,
            'atribuida': grupo.permissions.filter(id=permissao.id).exists()  # Verifica associação
        }
        for permissao in todas_permissoes
    ]

    if request.method == 'POST':
        for permissao in todas_permissoes:
            input_name = f"permissoes_{permissao.id}"
            atribuir = request.POST.get(input_name) == "sim"

            if atribuir and not grupo.permissions.filter(id=permissao.id).exists():
                grupo.permissions.add(permissao)  # Adiciona permissão
            elif not atribuir and grupo.permissions.filter(id=permissao.id).exists():
                grupo.permissions.remove(permissao)  # Remove permissão

        return redirect('listar_grupos')  # Redireciona após salvar

    return render(request, 'grupos_editar.html', {
        'grupo': grupo,
        'permissoes_detalhadas': permissoes_detalhadas
    })

import json  # Adicione esta linha


@csrf_exempt
def finalizar_titulo(request, titulo_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    if request.method == "POST":
        titulo = get_object_or_404(Titulo, id=titulo_id)
        titulo.ultima_acao = now().date()
        titulo.save()
        return JsonResponse({"status": "success", "message": "Título finalizado com sucesso!"})
    return JsonResponse({"status": "error", "message": "Método não permitido."}, status=405)

@csrf_exempt  # Permite AJAX, mas use o CSRF Token corretamente no cabeçalho

def atualizar_permissao(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            group_id = data.get('group_id')
            permission_id = data.get('permission_id')
            action = data.get('action')

            # Certifique-se de que os IDs são válidos
            grupo = Group.objects.get(id=group_id)
            permissao = Permission.objects.get(id=permission_id)

            if action == "sim":
                grupo.permissions.add(permissao)
            elif action == "nao":
                grupo.permissions.remove(permissao)
            else:
                return JsonResponse({"success": False, "error": "Ação inválida."})

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Método inválido."})





def permission_denied_view(request, exception):
    return render(request, '403.html', status=403)



def salvar_permissoes(request, grupo_id):
    grupo = Group.objects.get(id=grupo_id)
    if request.method == "POST":
        permissoes = Permission.objects.all()
        for permissao in permissoes:
            # Obtém valor do formulário
            valor = request.POST.get(f'permissoes_{permissao.id}', 'nao')
            if valor == "sim":
                grupo.permissions.add(permissao)  # Adiciona permissão
            else:
                grupo.permissions.remove(permissao)  # Remove permissão
    return redirect('listar_grupos')



def listar_permissoes_ptbr():
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    permissoes_por_modelo = {}
    permissoes = Permission.objects.select_related('content_type').all()

    for permissao in permissoes:
        modelo = permissao.content_type.model
        app_label = permissao.content_type.app_label
        nome = permissao.name

        descricao_traduzida = traduzir_permissao(nome)
        permissoes_por_modelo.setdefault(f"{app_label} - {modelo}", []).append(descricao_traduzida)

    return permissoes_por_modelo

def traduzir_permissao(permissao):
    # Traduções das permissões padrões do Django
    traducao = {
        'Can add': 'Pode adicionar',
        'Can change': 'Pode editar',
        'Can delete': 'Pode excluir',
        'Can view': 'Pode visualizar',
        # Adicione outras traduções personalizadas aqui se necessário
    }

    # Busca e substitui o padrão "Can <ação> <modelo>"
    for termo_en, termo_pt in traducao.items():
        if termo_en in permissao:
            return permissao.replace(termo_en, termo_pt)
    return permissao  # Retorna o original se não encontrar tradução
    


def listar_permissoes_view(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    permissoes = listar_permissoes_ptbr()
    return render(request, 'listar_permissoes.html', {'permissoes': permissoes})
    
    

def adicionar_usuario(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    groups = Group.objects.all()
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password_confirmation = request.POST.get('password_confirmation')
        group_name = request.POST.get('group')

        # Verificar se as senhas são iguais
        if password != password_confirmation:
            messages.error(request, "As senhas não conferem.")
            return render(request, 'usuarios_adicionar.html', {
                'groups': groups,
                'username': username,
                'email': email,
                'group_selected': group_name,
            })

        try:
            # Criar usuário e associar ao grupo
            user = User.objects.create_user(username=username, email=email, password=password)
            group = Group.objects.get(name=group_name)
            user.groups.add(group)
            messages.success(request, "Usuário criado com sucesso!")
            return redirect('listar_usuarios')
        except Exception as e:
            messages.error(request, f"Erro ao criar usuário: {e}")

    return render(request, 'usuarios_adicionar.html', {'groups': groups})







@lojista_login_required
def listar_usuarios(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    usuarios = User.objects.all()
    return render(request, 'usuarios_listar.html', {'usuarios': usuarios})

@lojista_login_required

def editar_usuario(request, user_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    user = get_object_or_404(User, id=user_id)
    groups = Group.objects.all()

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        password_confirmation = request.POST.get('password_confirmation')
        group_name = request.POST.get('group')

        # Verificar se as senhas são iguais
        if password and password != password_confirmation:
            messages.error(request, "As senhas não conferem.")
            return render(request, 'usuarios_editar.html', {
                'user': user,
                'groups': groups,
                'username': username,
                'email': email,
                'group_selected': group_name,
            })

        try:
            # Atualizar o usuário
            user.username = username
            user.email = email
            if password:
                user.set_password(password)
            user.save()

            # Atualizar o grupo
            user.groups.clear()
            group = Group.objects.get(name=group_name)
            user.groups.add(group)

            messages.success(request, "Usuário atualizado com sucesso!")
            return redirect('listar_usuarios')
        except Exception as e:
            messages.error(request, f"Erro ao editar usuário: {e}")

    # Carregar os grupos e dados do usuário no formulário
    return render(request, 'usuarios_editar.html', {
        'user': user,
        'groups': groups,
        'group_selected': user.groups.first().name if user.groups.exists() else None,
    })


@lojista_login_required

def excluir_usuario(request, user_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, f'Usuário {user.username} excluído com sucesso!')
    return redirect('listar_usuarios')




@lojista_login_required

def detalhar_parcela(request, parcela_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    parcela = get_object_or_404(Parcelamento, id=parcela_id)
    return render(request, 'detalhar_parcela.html', {'parcela': parcela})

# Devedores - Listar


@lojista_login_required
def devedores_listar(request):
    from django.contrib import messages
    from django.core.paginator import Paginator
    from django.db import connection
    from django.shortcuts import redirect, render

    # empresa da sessão
    empresa_id_sessao = request.session.get('empresa_id_sessao')
    if not empresa_id_sessao:
        messages.error(request, "Sessão expirada ou inválida. Por favor, faça login novamente.")
        return redirect('login')

    # filtros vindos do template
    q = request.GET.get('q', '').strip()          # <input name="q" ...>
    status = request.GET.get('status', '').strip()
    params = [empresa_id_sessao]

    # pega apenas o último título por devedor (evita linhas duplicadas)
    sql = """
        SELECT
            d.id              AS id,
            d.nome            AS nome,
            d.nome_fantasia   AS nome_fantasia,
            d.cpf             AS cpf,
            d.cnpj            AS cnpj,
            e.nome_fantasia   AS empresa,
            t.id              AS titulo_id,
            t.status_baixa    AS status_baixa
        FROM devedores d
        INNER JOIN core_empresa e
                ON d.empresa_id = e.id
        LEFT JOIN (
            SELECT tt.*
            FROM titulo tt
            INNER JOIN (
                SELECT devedor_id, MAX(id) AS max_id
                FROM titulo
                GROUP BY devedor_id
            ) ult
              ON ult.devedor_id = tt.devedor_id
             AND ult.max_id     = tt.id
        ) t
          ON t.devedor_id = d.id
        WHERE d.empresa_id = %s
    """

    if q:
        sql += """
            AND (
                 d.nome LIKE %s
              OR COALESCE(d.nome_fantasia,'') LIKE %s
              OR COALESCE(d.cpf,'')          LIKE %s
              OR COALESCE(d.cnpj,'')         LIKE %s
              OR e.nome_fantasia             LIKE %s
            )
        """
        like = f"%{q}%"
        params += [like, like, like, like, like]

    if status:
        # valores esperados: "Pendente", "Quitado", "Negociado"
        sql += " AND COALESCE(t.status_baixa,'') = %s"
        params.append(status)

    sql += " ORDER BY d.nome"

    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        cols = [c[0] for c in cursor.description]
        rows = cursor.fetchall()
        results = [dict(zip(cols, r)) for r in rows]

    paginator = Paginator(results, 10)  # 10 por página
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        "page_obj": page_obj,
        "query": q,
        "status": status,
    }
    return render(request, "devedores_listar.html", context)





@lojista_login_required
def baixar_modelo_devedor(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    # Consulta as core_empresa.e tipos de documentos no banco
    empresas = Empresa.objects.filter(id=empresa_id_sessao).values_list("nome_fantasia", flat=True)  # Aplica o filtro
    tipos_doc = TipoDocTitulo.objects.values_list("id", "name")  # Supondo que essa seja a tabela de tipos de documentos

    # Cria um Workbook com duas abas
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Devedores e Títulos"

    # Defina os campos do modelo
    campos = [
        "empresa_nome_fantasia", "tipo_pessoa", "cpf", "cnpj", "nome", "nome_mae", "rg",
        "razao_social", "nome_fantasia", "nome_socio", "telefone", "telefone1", "telefone2",
        "telefone3", "telefone4", "telefone5", "telefone6", "telefone7", "telefone8",
        "telefone9", "telefone10", "observacao", "cep", "endereco", "bairro", "uf", "cidade", "email1",
        # Campos de Título
        "num_titulo", "dataEmissao", "dataVencimento", "valor", "tipo_doc_id"
    ]
    ws1.append(campos)

    # Adiciona a aba com as empresas
    ws2 = wb.create_sheet("Empresas")
    for empresa in empresas:
        ws2.append([empresa])

    # Adiciona uma aba com os tipos de documento
    ws3 = wb.create_sheet("TiposDoc")
    for tipo in tipos_doc:
        ws3.append([f"{tipo[0]} - {tipo[1]}"])  # Exibe o ID e o nome do documento

    # Adiciona uma aba para as opções de tipo_pessoa
    ws4 = wb.create_sheet("Opções")
    ws4.append(["Tipo Pessoa"])
    ws4.append(["Física"])
    ws4.append(["Jurídica"])

    # Cria validação de dados (drop-down) para a coluna `empresa_nome_fantasia`
    empresa_dv = DataValidation(
        type="list",
        formula1=f"'Empresas'!$A$1:$A${len(empresas)}",
        allow_blank=False,
        showErrorMessage=True
    )
    ws1.add_data_validation(empresa_dv)
    for row in range(2, 1002):
        empresa_dv.add(ws1[f"A{row}"])  # Aplica à coluna `empresa_nome_fantasia`

    # Cria validação de dados para `tipo_pessoa`
    tipo_pessoa_dv = DataValidation(
        type="list",
        formula1=f"'Opções'!$A$2:$A$3",
        allow_blank=False,
        showErrorMessage=True
    )
    ws1.add_data_validation(tipo_pessoa_dv)
    for row in range(2, 1002):
        tipo_pessoa_dv.add(ws1[f"B{row}"])  # Aplica à coluna `tipo_pessoa`

    # Cria validação de dados para `tipo_doc_id`
    tipo_doc_dv = DataValidation(
        type="list",
        formula1=f"'TiposDoc'!$A$1:$A${len(tipos_doc)}",
        allow_blank=False,
        showErrorMessage=True
    )
    ws1.add_data_validation(tipo_doc_dv)
    for row in range(2, 1002):
        tipo_doc_dv.add(ws1[f"AG{row}"])  # Aplica à coluna `tipo_doc_id` (coluna AG)

    # Gera o arquivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="modelo_importacao_devedor_titulo.xlsx"'
    wb.save(response)
    return response





# views.py  (app_lojista)  — importar_devedor com mapeamento de tipo_doc_id
import os, re, math, unicodedata, tempfile, logging
from datetime import datetime
import pandas as pd

from django.contrib import messages
from django.shortcuts import redirect
from django.views.decorators.http import require_POST

logger = logging.getLogger(__name__)

# ==== mapeamento de nomes -> IDs (ajuste os números conforme seu banco) ====
TIPO_DOC_MAP = {
    "boleto": 1, "bl": 1,
    "cheque": 2, "ch": 2,
    "duplicata": 3, "dup": 3,
    "nota promissoria": 4, "promissoria": 4, "np": 4,
    "contrato": 5,
    "carne": 6, "carne de pagamento": 6, "carnê": 6,
    "fatura": 7,
    "outros": 9, "outro": 9,
}

def _norm_text(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()
    return s

def resolve_tipo_doc_id(val):
    """
    Aceita:
      - 2
      - "2 - Cheque"
      - "Cheque", "nota promissoria", "duplicata", etc.
    Retorna um int ou None se não reconhecer.
    """
    if val is None:
        return None
    # se vier NaN do pandas
    if isinstance(val, float):
        try:
            if math.isnan(val):
                return None
        except Exception:
            pass

    s = str(val).strip()
    # pega número inicial, se existir: "2 - Cheque" -> 2
    m = re.match(r"^\s*(\d+)\b", s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            pass

    key = _norm_text(s)
    return TIPO_DOC_MAP.get(key)

def normaliza_tipo_pessoa(v):
    if v is None:
        return 'F'
    try:
        if isinstance(v, float) and math.isnan(v):
            return 'F'
    except Exception:
        pass
    s = _norm_text(v)
    if "juridica" in s or s in {"j", "pj", "2"}:
        return "J"
    return "F"

def parse_date(val):
    if val is None:
        return None
    try:
        if isinstance(val, float) and math.isnan(val):
            return None
    except Exception:
        pass
    d = pd.to_datetime(val, dayfirst=True, errors="coerce")
    try:
        return d.date() if not pd.isna(d) else None
    except Exception:
        return None

def only_digits(s, max_len=None):
    if s is None:
        return None
    ss = re.sub(r"\D+", "", str(s))
    if max_len:
        ss = ss[:max_len]
    return ss or None

@require_POST
@lojista_login_required
def importar_devedor(request):
    f = request.FILES.get("arquivo")
    if not f:
        messages.error(request, "Nenhum arquivo foi enviado.")
        return redirect("listar_devedores")

    ext = os.path.splitext(f.name)[1].lower()
    if ext not in (".xlsx", ".xls", ".csv"):
        messages.error(request, "Formato não suportado. Envie .xlsx, .xls ou .csv.")
        return redirect("listar_devedores")
    if f.size > 20 * 1024 * 1024:
        messages.error(request, "Arquivo muito grande. Limite: 20 MB.")
        return redirect("listar_devedores")

    tmp_path = None
    try:
        # salva temporário
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            for chunk in f.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        # lê planilha
        try:
            if ext == ".csv":
                df = pd.read_csv(tmp_path)
            elif ext == ".xls":
                df = pd.read_excel(tmp_path, engine="xlrd")
            else:
                df = pd.read_excel(tmp_path, engine="openpyxl")
        except Exception as e:
            messages.error(request, f"Não foi possível ler a planilha: {e}")
            return redirect("listar_devedores")

        required = ["empresa_nome_fantasia", "tipo_pessoa", "nome", "num_titulo", "valor"]
        missing = [c for c in required if c not in df.columns]
        if missing:
            messages.error(request, "Planilha inválida. Faltam colunas: " + ", ".join(missing))
            return redirect("listar_devedores")

        criados, erros = 0, 0

        for i, row in df.iterrows():
            linha = i + 2  # header = linha 1

            # Empresa
            empresa_nome = str(row.get("empresa_nome_fantasia") or "").strip()
            empresa = Empresa.objects.filter(nome_fantasia=empresa_nome).first()
            if not empresa:
                erros += 1
                messages.error(request, f"Linha {linha}: Empresa não encontrada: {empresa_nome}")
                continue

            # Datas
            data_emissao    = parse_date(row.get("dataEmissao"))
            data_vencimento = parse_date(row.get("dataVencimento"))

            # Devedor
            tp   = normaliza_tipo_pessoa(row.get("tipo_pessoa"))
            cpf  = only_digits(row.get("cpf"), 11)
            cnpj = only_digits(row.get("cnpj"), 14)

            tel_fields = {}
            for k in ["telefone","telefone1","telefone2","telefone3","telefone4","telefone5",
                      "telefone6","telefone7","telefone8","telefone9","telefone10"]:
                tel_fields[k] = only_digits(row.get(k), 20)

            try:
                devedor, _ = Devedor.objects.get_or_create(
                    empresa=empresa,
                    tipo_pessoa=tp,
                    cpf=cpf,
                    cnpj=cnpj,
                    defaults={
                        "nome": row.get("nome"),
                        "nome_mae": row.get("nome_mae"),
                        "rg": row.get("rg"),
                        "razao_social": row.get("razao_social"),
                        "nome_fantasia": row.get("nome_fantasia"),
                        "nome_socio": row.get("nome_socio"),
                        **tel_fields,
                        "observacao": row.get("observacao"),
                        "cep": only_digits(row.get("cep"), 8),
                        "endereco": row.get("endereco"),
                        "bairro": row.get("bairro"),
                        "uf": str(row.get("uf") or "")[:2],
                        "cidade": row.get("cidade"),
                        "email1": row.get("email1"),
                    },
                )
            except Exception as e:
                erros += 1
                messages.error(request, f"Linha {linha}: erro ao criar/obter devedor: {e}")
                continue

            # Título
            num_titulo = row.get("num_titulo")
            if pd.isna(num_titulo) or str(num_titulo).strip() == "":
                erros += 1
                messages.error(request, f"Linha {linha}: num_titulo vazio.")
                continue

            try:
                valor = float(row.get("valor") or 0)
            except Exception:
                erros += 1
                messages.error(request, f"Linha {linha}: valor inválido ({row.get('valor')}).")
                continue

            # --> AQUI: resolve tipo_doc_id de forma flexível
            bruto_tipo = row.get("tipo_doc_id") or row.get("tipo_doc") or row.get("documento")
            tipo_doc_id = resolve_tipo_doc_id(bruto_tipo)
            if not tipo_doc_id:
                erros += 1
                keys = ", ".join(sorted(TIPO_DOC_MAP.keys()))
                messages.error(
                    request,
                    f"Linha {linha}: tipo_doc_id inválido ({bruto_tipo}). "
                    f"Use número (ex.: '2' ou '2 - Cheque') ou um dos nomes: {keys}."
                )
                continue

            try:
                Titulo.objects.create(
                    devedor=devedor,
                    num_titulo=str(num_titulo),
                    dataEmissao=data_emissao,
                    dataVencimento=data_vencimento,
                    valor=valor,
                    tipo_doc_id=tipo_doc_id,
                    acordoComfirmed=0,
                    renegociado=0,
                )
                criados += 1
            except Exception as e:
                erros += 1
                messages.error(request, f"Linha {linha}: erro ao criar título ({num_titulo}): {e}")
                continue

        messages.success(request, f"Importação finalizada: {criados} título(s) criado(s), {erros} erro(s).")
        return redirect("listar_devedores")

    except Exception as e:
        messages.error(request, f"Erro geral durante a importação: {e}")
        return redirect("listar_devedores")

    finally:
        if tmp_path:
            try: os.remove(tmp_path)
            except Exception: pass





def agendamentos_cadastrar(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')

    if request.method == 'POST':
        try:
            devedor_id = request.POST.get('devedor_id')
            empresa_id = request.POST.get('empresa_id')
            telefone = request.POST.get('telefone')  # Captura o telefone
            data_abertura = request.POST.get('data_abertura')
            data_retorno = request.POST.get('data_retorno')
            assunto = request.POST.get('assunto')
            operador = request.POST.get('operador')

            # Validação básica
            if not devedor_id or not empresa_id:
                messages.error(request, "Devedor e Empresa são obrigatórios.")
                return redirect('agendamentos_cadastrar')

            # Atualiza o telefone no devedor, se aplicável
            devedor = Devedor.objects.filter(empresa_id=empresa_id_sessao).get(id=devedor_id)
            if telefone:
                devedor.telefone = telefone
                devedor.save()

            # Criação do agendamento
            Agendamento.objects.create(
                devedor=devedor,
                empresa_id=empresa_id,
                data_abertura=data_abertura,
                data_retorno=data_retorno,
                assunto=assunto,
                operador=operador
            )

            messages.success(request, "Agendamento criado com sucesso.")
            return redirect('agendamentos_listar')

        except Exception as e:
            messages.error(request, f"Erro ao criar agendamento: {e}")

    devedores = Devedor.objects.filter(empresa_id=empresa_id_sessao).values(
        'id', 'nome', 'telefone'
    )
    return render(request, 'agendamentos_criar.html', {'devedores': devedores})



# Agendamentos - Listar
@lojista_login_required
def agendamentos_listar(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    return render(request, 'agendamentos_listar.html')

# Página de Login

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if not username or not password:
            messages.error(request, 'Preencha todos os campos.')
            return render(request, 'login.html')

        try:
            user = UserLojista.objects.get(email=username)
            stored_hashed_password = user.password.replace("$2y$", "$2b$", 1)

            if bcrypt.checkpw(password.encode('utf-8'), stored_hashed_password.encode('utf-8')):
                # Busca o nome fantasia da empresa associada
                empresa = Empresa.objects.filter(id=user.empresa_id).first()
                if empresa:
                    # Salva os dados na sessão
                    request.session['user_id'] = user.id
                    request.session['user_name'] = user.name
                    request.session['empresa_id_sessao'] = user.empresa_id
                    request.session['nome_fantasia_sessao'] = empresa.nome_fantasia  # Adiciona o nome fantasia na sessão
                    request.session['is_authenticated'] = True
                    return redirect('dashboard')
                else:
                    messages.error(request, 'Empresa associada ao usuário não encontrada.')
            else:
                messages.error(request, 'Senha incorreta.')

        except UserLojista.DoesNotExist:
            messages.error(request, 'Usuário não encontrado.')

    return render(request, 'login.html')





@lojista_login_required
def listar_devedores(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão

    # Obtém o termo de pesquisa e status de filtro (se fornecidos)
    search_query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')

    # Mapeamento de status para os valores numéricos no banco
    status_map = {
        "Pendente": 0,  # Considera apenas 0 para pendentes
        "Quitado": 2,
        "Negociado": 3,
    }

    # Consulta SQL base ajustada para incluir o filtro de empresa_id_sessao
    query = """
        SELECT 
            MAX(titulo.statusBaixa) AS status_baixa,
            devedores.id AS devedor_id, 
            devedores.nome AS devedor_nome, 
            devedores.cpf AS devedor_cpf, 
            devedores.cnpj AS devedor_cnpj, 
            core_empresa.nome_fantasia AS empresa_nome, 
            COUNT(titulo.id) AS quantidade_titulos,
            devedores.nome_fantasia,
            devedores.razao_social            
        FROM 
            devedores
        INNER JOIN titulo ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE core_empresa.id = %s
    """ % (empresa_id_sessao,)  # Usar parametrização para evitar SQL Injection

    # Adiciona condições de pesquisa à consulta SQL, se aplicável
    if search_query:
        query += f"""
            AND (
                devedores.nome LIKE '%{search_query}%'
                OR devedores.cpf LIKE '%{search_query}%'
                OR devedores.cnpj LIKE '%{search_query}%'
                OR devedores.telefone LIKE '%{search_query}%'
                OR core_empresa.nome_fantasia LIKE '%{search_query}%'
                OR devedores.nome_fantasia LIKE '%{search_query}%'
                OR devedores.razao_social LIKE '%{search_query}%'
            )
        """

    # Adiciona filtro de status, se aplicável
    if status_filter:
        status_value = status_map.get(status_filter, None)
        if status_value is not None:
            query += f" AND titulo.statusBaixa = {status_value}"

    # Agrupamento e ordenação
    query += """
        GROUP BY devedores.id, devedores.nome, devedores.cpf, devedores.cnpj, core_empresa.nome_fantasia
        ORDER BY devedores.id
    """

    # Executa a consulta diretamente no banco
    with connection.cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()

    # Mapeia os resultados para um dicionário
    devedores = [
        {
            "id": row[1],
            "nome": row[2],
            "cpf": row[3],
            "cnpj": row[4] if row[4] else "Não informado",
            "empresa": row[5],
            "quantidade_titulos": row[6],  # Número total de títulos
            "nome_fantasia": row[7],
            "razao_social": row[8],
            "display_nome": (row[2] or row[7] or row[8] or f"Devedor {row[1]}").strip(),
            "status_baixa": {
                0: "Pendente",
                2: "Quitado",
                3: "Negociado",
                None: "Pendente",
            }.get(row[0], "Desconhecido"),
        }
        for row in rows
    ]

    total_devedores = len(devedores)
    total_pendentes = sum(1 for d in devedores if d["status_baixa"] == "Pendente")
    total_quitados = sum(1 for d in devedores if d["status_baixa"] == "Quitado")
    total_negociados = sum(1 for d in devedores if d["status_baixa"] == "Negociado")

    # Configura a paginação (10 itens por página)
    paginator = Paginator(devedores, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': search_query,
        'status': status_filter,
        'total_devedores': total_devedores,
        'total_pendentes': total_pendentes,
        'total_quitados': total_quitados,
        'total_negociados': total_negociados,
        'status_options': ["Pendente", "Quitado", "Negociado"],
    }

    # Renderiza o template com os dados e os filtros
    return render(request, 'devedores_listar.html', context)



# Adicionar Devedor
# Adicionar Devedor
@lojista_login_required    
def adicionar_devedor(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao = None
    if empresa_id_sessao:
        empresa_sessao = Empresa.objects.filter(id=empresa_id_sessao).first()  # Busca a empresa da sessão

    empresas = Empresa.objects.all()  # Para preencher o campo empresa
    fields = [
        'cpf', 'cnpj', 'nome', 'nome_mae', 'rg', 'razao_social',
        'nome_fantasia', 'nome_socio', 'cpf_socio', 'rg_socio', 'telefone',
        'telefone1', 'telefone2', 'telefone3', 'telefone4', 'telefone5',
        'telefone6', 'telefone7', 'telefone8', 'telefone9', 'telefone10',
        'cep', 'endereco', 'bairro', 'uf', 'cidade', 'email1', 'email2',
        'observacao'
    ]

    initial_data = {field: '' for field in fields}
    initial_tipo = 'F'

    if request.method == 'POST':
        data = request.POST
        empresa = Empresa.objects.get(id=data['empresa_id'])
        devedor_data = {field: data.get(field) for field in fields}  # Captura todos os campos dinamicamente
        devedor = Devedor.objects.create(
            empresa=empresa,
            tipo_pessoa=data['tipo_pessoa'],
            **devedor_data,  # Descompacta os campos capturados
        )
        messages.success(request, 'Devedor adicionado com sucesso.')

        # Redireciona para a página de adicionar título para o devedor recém-criado
        return redirect('adicionar_titulo_pg_devedor', devedor_id=devedor.id)
    else:
        data = request.GET

    for field in fields:
        initial_data[field] = data.get(field, '')
    initial_tipo = data.get('tipo_pessoa', 'F') or 'F'
    initial_data['tipo_pessoa'] = initial_tipo
    initial = SimpleNamespace(**initial_data)

    return render(request, 'devedores_adicionar.html', {
        'empresas': empresas,
        'fields': fields,
        'empresa_sessao': empresa_sessao,  # Passa a empresa da sessão
        'initial': initial,
        'initial_tipo': initial_tipo,
    })


from decimal import Decimal, InvalidOperation
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.urls import reverse

@lojista_login_required
def adicionar_titulo_pg_devedor(request, devedor_id):
    # Inicializar variáveis
    devedor = None
    empresa_selecionada = None
    tipos_docs = []
    
    try:
        empresa_id_sessao = request.session.get('empresa_id_sessao')
        if not empresa_id_sessao:
            messages.error(request, 'Empresa não selecionada na sessão.')
            return redirect('login')
        
        devedor = get_object_or_404(Devedor, id=devedor_id)
        
        # Verifica se o devedor pertence à empresa da sessão
        if devedor.empresa_id != empresa_id_sessao:
            messages.error(request, 'Você não tem permissão para adicionar título para este devedor.')
            return redirect('listar_devedores')
        
        empresa_selecionada = Empresa.objects.filter(id=empresa_id_sessao).first()
        if not empresa_selecionada:
            messages.error(request, 'Empresa não encontrada.')
            return redirect('login')
        
        tipos_docs = TipoDocTitulo.objects.all()
    except Http404:
        messages.error(request, 'Devedor não encontrado.')
        return redirect('listar_devedores')
    except Exception as e:
        logger.error(f"Erro ao carregar página de adicionar título: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        messages.error(request, 'Erro ao carregar página.')
        return redirect('listar_devedores')

    if request.method == 'POST':
        d = request.POST

        # Empresa
        empresa_id = d.get('empresa_id') or (empresa_selecionada.id if empresa_selecionada else None)
        if not empresa_id:
            messages.error(request, 'Selecione a empresa.')
            return render(request, 'titulos_adicionar_pg_devedor.html', {
                'devedor': devedor, 'tipos_docs': tipos_docs, 'empresa_selecionada': empresa_selecionada
            })
        empresa = get_object_or_404(Empresa, id=empresa_id)

        # Tipo doc
        tipo_doc_id_str = (d.get('tipo_doc_id') or '').strip()
        if not tipo_doc_id_str:
            messages.error(request, 'Selecione o tipo de documento.')
            return render(request, 'titulos_adicionar_pg_devedor.html', {
                'devedor': devedor, 'tipos_docs': tipos_docs, 'empresa_selecionada': empresa_selecionada
            })
        
        try:
            tipo_doc = get_object_or_404(TipoDocTitulo, id=int(tipo_doc_id_str))
        except (ValueError, TypeError):
            messages.error(request, 'Tipo de documento inválido.')
            return render(request, 'titulos_adicionar_pg_devedor.html', {
                'devedor': devedor, 'tipos_docs': tipos_docs, 'empresa_selecionada': empresa_selecionada
            })

        # Campos
        num_titulo_str = (d.get('num_titulo') or '').strip()
        data_emissao = (d.get('data_emissao') or '').strip()
        data_venc = (d.get('data_vencimento') or '').strip()
        valor_str = (d.get('valor') or '').strip()
        status_str = (d.get('status_baixa') or '0').strip()

        if not (num_titulo_str and data_emissao and data_venc and valor_str):
            messages.error(request, 'Preencha número, emissão, vencimento e valor.')
            return render(request, 'titulos_adicionar_pg_devedor.html', {
                'devedor': devedor, 'tipos_docs': tipos_docs, 'empresa_selecionada': empresa_selecionada
            })
        
        # Validar e converter num_titulo para inteiro
        try:
            num_titulo = int(num_titulo_str)
        except ValueError:
            messages.error(request, 'Número do título deve ser um valor numérico válido.')
            return render(request, 'titulos_adicionar_pg_devedor.html', {
                'devedor': devedor, 'tipos_docs': tipos_docs, 'empresa_selecionada': empresa_selecionada
            })

        # Datas
        try:
            data_emissao = datetime.strptime(data_emissao, '%Y-%m-%d').date()
            data_venc = datetime.strptime(data_venc, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Datas inválidas. Use AAAA-MM-DD.')
            return render(request, 'titulos_adicionar_pg_devedor.html', {
                'devedor': devedor, 'tipos_docs': tipos_docs, 'empresa_selecionada': empresa_selecionada
            })

        # Valor - converter para float (o modelo usa FloatField)
        try:
            valor_decimal = Decimal(valor_str.replace('.', '').replace(',', '.'))
            valor = float(valor_decimal)
        except (InvalidOperation, AttributeError, ValueError):
            messages.error(request, 'Valor inválido. Use 1.234,56.')
            return render(request, 'titulos_adicionar_pg_devedor.html', {
                'devedor': devedor, 'tipos_docs': tipos_docs, 'empresa_selecionada': empresa_selecionada
            })

        try:
            status_baixa = int(status_str)
        except ValueError:
            status_baixa = 0

        # CREATE — sem statusBaixaGeral
        try:
            Titulo.objects.create(
                empresa=empresa,
                devedor=devedor,
                num_titulo=num_titulo,
                valor=valor,
                dataVencimento=data_venc,
                dataEmissao=data_emissao,
                tipo_doc=tipo_doc,
                statusBaixa=status_baixa,
                idTituloRef=None,
                renegociado=0,
            )
            messages.success(request, 'Título adicionado com sucesso.')
            next_url = request.GET.get('next') or reverse('listar_titulos_por_devedor', args=[devedor.id])
            return redirect(next_url)
        except Exception as e:
            logger.error(f"Erro ao criar título: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            messages.error(request, f'Erro ao salvar título: {str(e)}')
            return render(request, 'titulos_adicionar_pg_devedor.html', {
                'devedor': devedor, 'tipos_docs': tipos_docs, 'empresa_selecionada': empresa_selecionada
            })

    # Verificar se todas as variáveis necessárias estão definidas
    if not devedor or not empresa_selecionada:
        messages.error(request, 'Erro ao carregar dados do formulário.')
        return redirect('listar_devedores')
    
    try:
        return render(request, 'titulos_adicionar_pg_devedor.html', {
            'devedor': devedor,
            'tipos_docs': tipos_docs,
            'empresa_selecionada': empresa_selecionada,
        })
    except Exception as e:
        logger.error(f"Erro ao renderizar template: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        messages.error(request, 'Erro ao carregar formulário.')
        return redirect('listar_devedores')







@lojista_login_required
def editar_devedor(request, id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pega o empresa_id da sessão
    devedor = get_object_or_404(Devedor, id=id)

    # Verifica se o devedor pertence à empresa na sessão
    if devedor.empresa_id != empresa_id_sessao:
        # Se não pertence, retorna um erro 403 Forbidden
        return HttpResponseForbidden("Você não tem permissão para editar devedores de outra empresa.")

    empresas = Empresa.objects.all()  # Para popular o dropdown de empresas
    telefones_range = range(1, 11)  # Lista de números para os telefones

    if request.method == 'POST':
        empresa_id = request.POST.get('empresa_id')
        tipo_pessoa = request.POST.get('tipo_pessoa')
        
        # Atualizar os campos individuais
        devedor.nome = request.POST.get('nome', '')
        devedor.cpf = request.POST.get('cpf', '')
        devedor.cnpj = request.POST.get('cnpj', '')
        devedor.razao_social = request.POST.get('razao_social', '')
        devedor.nome_fantasia = request.POST.get('nome_fantasia', '')
        devedor.cep = request.POST.get('cep', '')
        devedor.endereco = request.POST.get('endereco', '')
        devedor.bairro = request.POST.get('bairro', '')
        devedor.uf = request.POST.get('uf', '')
        devedor.cidade = request.POST.get('cidade', '')

        # Telefones
        for i in telefones_range:
            setattr(devedor, f'telefone{i}', request.POST.get(f'telefone{i}', ''))

        devedor.observacao = request.POST.get('observacao', '')

        # Atualizar empresa associada
        if empresa_id and empresa_id == empresa_id_sessao:
            devedor.empresa_id = empresa_id
        else:
            return HttpResponseForbidden("Você não tem permissão para alterar o devedor para outra empresa.")

        # Atualizar tipo de pessoa
        devedor.tipo_pessoa = tipo_pessoa

        # Salvar as alterações no banco de dados
        devedor.save()
        messages.success(request, "Devedor atualizado com sucesso!")
        return redirect('listar_devedores')

    # Passar os dados do devedor para o template
    return render(request, 'devedores_editar.html', {
        'devedor': devedor,
        'empresas': empresas,
        'telefones_range': telefones_range,
    })


@lojista_login_required
# Excluir Devedor
def excluir_devedor(request, id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    devedor = get_object_or_404(Devedor, id=id)
    if request.method == 'POST':
        devedor.delete()
        messages.success(request, 'Devedor excluído com sucesso.')
        return redirect('listar_devedores')
    return render(request, 'devedores_excluir.html', {'devedor': devedor})
    




@lojista_login_required
def titulos_listar(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    # Obtém os parâmetros de busca e filtro
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')

    # Mapeamento de status
    status_map = {
        "Pendente": (0, "NULL"),  # Considera 0 e NULL para pendentes
        "Quitado": 2,
        "Negociado": 3,
    }

    # Construindo a consulta SQL base
    query_sql = """
        SELECT 
            titulo.id, 
            core_empresa.razao_social, 
            titulo.num_titulo, 
            titulo.valor, 
            titulo.dataVencimento AS data_vencimento,
            titulo.data_baixa, 
            titulo.statusBaixa AS status_baixa,
            devedores.nome AS devedor_nome,
            devedores.cpf AS devedor_cpf,
            core_empresa.nome_fantasia,            
            devedores.cnpj
        FROM 
            titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE 1=1
    """

    # Adiciona condição de busca, se aplicável
    query_params = []
    if query:
        query_sql += """
            AND (
                titulo.num_titulo LIKE %s OR
                devedores.nome LIKE %s OR
                devedores.cpf LIKE %s OR
                core_empresa.razao_social LIKE %s OR
                core_empresa.nome_fantasia LIKE %s OR
                devedores.cnpj LIKE %s
            )
        """
        query_params = [f'%{query}%'] * 6

    # Adiciona condição de filtro por status, se aplicável
    if status_filter:
        if status_filter == "Pendente":
            query_sql += " AND (titulo.statusBaixa = 0 OR titulo.statusBaixa IS NULL)"
        else:
            query_sql += " AND titulo.statusBaixa = %s"
            query_params.append(status_map[status_filter])

    # Ordena resultados
    query_sql += " ORDER BY titulo.id DESC"

    # Executa a consulta
    with connection.cursor() as cursor:
        cursor.execute(query_sql, query_params)
        rows = cursor.fetchall()

    # Mapeia os resultados
    titulos = [
        {
            'id': row[0],
            'razao_social': row[1],
            'num_titulo': row[2],
            'valor': row[3],
            'data_vencimento': row[4],
            'data_baixa': row[5],
            'status_baixa': row[6],
            'devedor_nome': row[7],
            'devedor_cpf': row[8],  # Adicionando o CPF
            'nome_fantasia': row[9],            
            'cnpj': row[10],
        }
        for row in rows
    ]

    # Configura paginação
    paginator = Paginator(titulos, 30)  # Limita 30 registros por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'titulos_listar.html', {
        'page_obj': page_obj, 
        'query': query, 
        'status': status_filter
    })



@lojista_login_required
def listar_titulos_por_devedor(request, devedor_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    """
    Lista os títulos associados a um devedor específico.
    """
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                titulo.id AS titulo_id,
                core_empresa.razao_social AS empresa_razao_social,
                titulo.num_titulo AS numero_titulo,
                titulo.valor AS valor_titulo,
                titulo.dataVencimento AS data_vencimento,
                titulo.data_baixa AS data_baixa,
                titulo.statusBaixa AS status_baixa,
                devedores.nome AS nome_devedor,
                titulo.valorRecebido,
                titulo.forma_pag_Id,
                titulo.idTituloRef,
                titulo.juros,
                titulo.dias_atraso
            FROM 
                titulo
            INNER JOIN devedores ON titulo.devedor_id = devedores.id
            INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
            WHERE devedores.id = %s
        """, [devedor_id])
        rows = cursor.fetchall()

    # Mapeamento de formas de pagamento
    forma_pagamento_map = {
        0: "Pix",
        1: "Dinheiro",
        2: "Cartão de Débito",
        3: "Cartão de Crédito",
        4: "Cheque",
        5: "Depósito em Conta",
        6: "Pagamento na Loja",
        7: "Boleto Bancário",
        8: "Duplicata",
    }

    # Inicialização de variáveis
    titulos_principais = []
    titulos_entrada = []
    entrada_ids = set()

    total_quitado = 0
    total_negociado = 0
    total_pendente = 0

    for row in rows:
        valor = row[3] or 0
        valor_recebido = row[8] or 0
        status_baixa = row[6] or 0  # Trata como 0 se for None
        titulo_dict = {
            'id': row[0],
            'razao_social': row[1],
            'num_titulo': row[2],
            'valor': valor,
            'data_vencimento': row[4],
            'data_baixa': row[5],
            'status_baixa': status_baixa,
            'devedor_nome': row[7],
            'valor_recebido': valor_recebido,
            'forma_pagamento': forma_pagamento_map.get(row[9], "Não definido"),
            'juros': row[11],
            'dias_atraso': row[12],
        }

        # Soma para totais
        if status_baixa == 2:
            total_quitado += valor_recebido
        elif status_baixa == 3:
            total_negociado += valor
        elif status_baixa == 0 or status_baixa is None:
            total_pendente += valor

        # Adiciona a entrada se for o caso
        if row[10] is None and status_baixa > 1:
            titulos_entrada.append(titulo_dict)
            entrada_ids.add(row[0])

        titulos_principais.append(titulo_dict)

    today = date.today()

    # Passa os valores para o contexto
    return render(request, 'titulos_listar_por_devedor.html', {
        'titulos': titulos_principais,
        'titulos_entrada': titulos_entrada,
        'entrada_ids': entrada_ids,
        'devedor_id': devedor_id,
        'total_quitado': total_quitado,
        'total_negociado': total_negociado,
        'total_pendente': total_pendente,
        'today': today,
    })





def negociacao_devedor(request, devedor_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    devedor = get_object_or_404(Devedor, id=devedor_id)

    # Calcular os totais
    total_quitado = Titulo.objects.filter(devedor=devedor, statusBaixa=2).aggregate(Sum('valorRecebido'))['valorRecebido__sum'] or 0
    total_negociado = Titulo.objects.filter(devedor=devedor, statusBaixa=3).aggregate(Sum('valor'))['valor__sum'] or 0
    total_pendente = Titulo.objects.filter(devedor=devedor, statusBaixa__in=[0, None]).aggregate(Sum('valor'))['valor__sum'] or 0

    # Debugging - Printando no console
    print(f"Total Quitado: {total_quitado}")
    print(f"Total Negociado: {total_negociado}")
    print(f"Total Pendente: {total_pendente}")

    # Passando o contexto para o template
    context = {
        'devedor': devedor,
        'total_quitado': total_quitado,
        'total_negociado': total_negociado,
        'total_pendente': total_pendente,
    }
    return render(request, 'negociacao_devedor.html', context)



from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

@lojista_login_required
def adicionar_titulo(request):
    # empresa vinda da sessão (quando o lojista está logado)
    empresa_id_sessao = request.session.get('empresa_id_sessao') or request.session.get('empresa_sessao_id')
    empresa_sessao = Empresa.objects.filter(id=empresa_id_sessao).first() if empresa_id_sessao else None

    if request.method == 'POST':
        data = request.POST

        # Se há empresa na sessão, usa ela; senão, pega do form
        empresa_id = empresa_sessao.id if empresa_sessao else data.get('empresa_id')
        empresa = get_object_or_404(Empresa, id=empresa_id)

        # Devedor precisa pertencer à mesma empresa
        devedor = get_object_or_404(Devedor, id=data.get('devedor_id'), empresa_id=empresa.id)

        # Valor → Decimal (aceita "1.234,56")
        raw_valor = (data.get('valor') or '0').replace('.', '').replace(',', '.')
        try:
            valor = Decimal(raw_valor)
        except Exception:
            messages.error(request, 'Valor inválido.')
            return redirect(request.path)

        # Datas podem vir com nomes diferentes no form
        data_venc = data.get('data_vencimento') or data.get('dataVencimento')
        data_emis = data.get('data_emissao') or data.get('dataEmissao')

        # Status (0 pendente, 2 quitado, 3 negociado)
        status = int(data.get('status_baixa') or 0)

        Titulo.objects.create(
            empresa=empresa,
            devedor=devedor,
            num_titulo=(data.get('num_titulo') or '').strip(),
            valor=valor,
            dataVencimento=data_venc,      # <- nome do campo no modelo
            dataEmissao=data_emis,         # <- se existir no seu modelo
            statusBaixa=status,
            statusBaixaGeral=status,       # <- MUITO IMPORTANTE p/ ficar igual ao Admin
            idTituloRef=None,              # <- evita ser marcado como "filho" de parcelamento
            renegociado=0,
        )

        messages.success(request, 'Título adicionado com sucesso.')
        # Melhor UX: voltar para os títulos do devedor criado
        return redirect('listar_titulos_por_devedor', devedor_id=devedor.id)

    # GET: listas para o formulário
    if empresa_sessao:
        # lojista só vê devedores da sua empresa
        devedores = Devedor.objects.filter(empresa=empresa_sessao).order_by('-id')
        empresas = [empresa_sessao]   # trava o select de empresa
    else:
        devedores = Devedor.objects.all().order_by('-id')
        empresas = Empresa.objects.all().order_by('nome_fantasia', 'razao_social')

    return render(request, 'titulos_adicionar.html', {
        'empresas': empresas,
        'devedores': devedores,
        'empresa_sessao': empresa_sessao,
    })



@lojista_login_required    
def editar_titulo(request, id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    titulo = get_object_or_404(Titulo, id=id)

    if request.method == 'POST':
        data = request.POST
        titulo.empresa = Empresa.objects.get(id=data['empresa_id'])
        titulo.devedor = Devedor.objects.get(id=data['devedor_id'])
        titulo.num_titulo = data['num_titulo']  # Atualizado para incluir o campo num_titulo
        
        # Trocar vírgula por ponto no valor
        valor = data['valor'].replace(',', '.')
        titulo.valor = valor

        titulo.dataVencimento = data['data_vencimento']
        titulo.tipo_doc_id = data['tipo_doc_id']  # Atualizado para incluir o campo tipo_doc_id
        titulo.statusBaixa = data.get('status_baixa', 0)
        titulo.save()

        messages.success(request, 'Título atualizado com sucesso.')
        return redirect('titulos_listar')

    # Convertendo dataVencimento para o formato YYYY-MM-DD
    if titulo.dataVencimento:
        titulo.dataVencimento = titulo.dataVencimento.strftime('%Y-%m-%d')

    core_empresa = Empresa.objects.all()
    devedores = Devedor.objects.all()
    tipos_docs = TipoDocTitulo.objects.all()  # Obter os tipos de documentos para o dropdown

    return render(request, 'titulos_editar.html', {
        'titulo': titulo,
        'core_empresa': core_empresa,
        'devedores': devedores,
        'tipos_docs': tipos_docs,  # Passar os tipos de documentos para o template
    })


    
    
@lojista_login_required    
def excluir_titulo(request, id):
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    titulo = get_object_or_404(Titulo, id=id)
    if request.method == 'POST':
        titulo.delete()
        messages.success(request, 'Título excluído com sucesso.')
        return redirect('titulos_listar')

    return render(request, 'titulos_excluir.html', {'titulo': titulo})







def validar_cnpj(cnpj):
    """
    Valida o formato e a estrutura do CNPJ.
    """
    cnpj = re.sub(r'\D', '', cnpj)  # Remove caracteres não numéricos
    if len(cnpj) != 14:
        return False
    
    # Validação básica para números sequenciais
    if cnpj in (c * 14 for c in "0123456789"):
        return False
    
    # Cálculo dos dígitos verificadores
    def calcular_digito(cnpj, peso):
        soma = sum(int(cnpj[i]) * peso[i] for i in range(len(peso)))
        resto = soma % 11
        return '0' if resto < 2 else str(11 - resto)
    
    primeiro_digito = calcular_digito(cnpj[:12], [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    segundo_digito = calcular_digito(cnpj[:13], [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2])
    
    return cnpj[12] == primeiro_digito and cnpj[13] == segundo_digito



def consultar_cnpj_view(request):
    cnpj = request.GET.get('cnpj', '').strip()
    if not cnpj:
        return JsonResponse({"erro": "CNPJ não fornecido"}, status=400)

    resultado = consultar_cnpj_via_scraping(cnpj)
    return JsonResponse(resultado)


def consultar_com_espera(cnpj):
    time.sleep(5)  # Espera 5 segundos entre as consultas
    return consultar_cnpj_via_scraping(cnpj)



@lojista_login_required
def realizar_acordo(request, titulo_id):
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    print(f"Tentando acessar o título com ID: {titulo_id}")

    try:
        titulo = Titulo.objects.get(id=titulo_id, devedor_id__isnull=False)
        print(f"Título encontrado: {titulo}")
    except Titulo.DoesNotExist:
        messages.error(request, "Título não encontrado ou não associado a um devedor.")
        return redirect('titulos_listar')

    diferenca_dias = (date.today() - titulo.dataVencimento).days if titulo.dataVencimento else 0
    juros_mensais = 0.08
    juros_totais = (titulo.valor * juros_mensais) * (diferenca_dias / 30) if diferenca_dias > 0 else 0
    
    # Formatar a data de vencimento para o formato dd/mm/yyyy
    data_vencimento_formatada = titulo.dataVencimento.strftime('%d/%m/%Y') if titulo.dataVencimento else ""
    
    if request.method == 'POST':
        data = request.POST
        try:
            entrada = float(data.get('entrada', 0))
            qtde_prc = int(data.get('qtde_prc', 0))
            valor_por_parcela = float(data.get('valor_por_parcela', 0))
            forma_pag_Id = int(data.get('forma_pag_Id', 0))
            venc_primeira_parcela = data.get('venc_primeira_parcela')

            if entrada <= 0 or qtde_prc <= 0 or valor_por_parcela <= 0:
                raise ValueError("Todos os valores devem ser maiores que zero.")

            # Atualizar título principal para refletir apenas o valor de entrada
            titulo.statusBaixa = 3
            titulo.valorRecebido = entrada
            titulo.total_acordo = entrada
            titulo.valor_parcela = None
            titulo.qtde_parcelas = None
            titulo.forma_pag_Id = None
            titulo.primeiro_vencimento = venc_primeira_parcela
            titulo.juros = juros_totais
            titulo.save()

            print(f"Título principal atualizado com sucesso: ID {titulo.id}")

            # Criar parcelas adicionais, incluindo a parcela inicial
            vencimento_inicial = datetime.strptime(venc_primeira_parcela, '%Y-%m-%d').date()
            for i in range(1, qtde_prc + 1):
                data_vencimento = vencimento_inicial + relativedelta(months=i - 1)

                nova_parcela = Titulo.objects.create(
                    idTituloRef=titulo.id,
                    num_titulo=titulo.num_titulo,
                    tipo_doc_id=titulo.tipo_doc_id,
                    dataEmissao=date.today(),
                    dataVencimento=data_vencimento,
                    dataVencimentoReal=data_vencimento,
                    dataVencimentoPrimeira=venc_primeira_parcela if i == 1 else None,
                    valor=valor_por_parcela,
                    qtde_parcelas=qtde_prc,
                    nPrc=i,  # Número da parcela
                    forma_pag_Id=None,
                    statusBaixa=3,  # Parcelas marcadas como negociadas
                    devedor_id=titulo.devedor_id,
                    renegociado=1,  # Parcelas são consideradas renegociadas
                )
                print(f"Parcela {i} criada com sucesso: ID {nova_parcela.id}")

            #messages.success(request, "Acordo realizado com sucesso!")
            return redirect('listar_titulos_por_devedor', titulo.devedor.id)

        except ValueError as e:
            print(f"Erro nos valores fornecidos: {e}")
            messages.error(request, f"Erro nos valores fornecidos: {e}")
        except Exception as e:
            print(f"Erro inesperado ao criar o acordo: {e}")
            messages.error(request, f"Erro inesperado: {e}")

    valor_total_com_juros = titulo.valor + juros_totais
    context = {
        'titulo': titulo,
        'juros_totais': juros_totais,
        'diferenca_dias': diferenca_dias,
        'valor_total_com_juros': valor_total_com_juros,
        'data_vencimento_formatada': data_vencimento_formatada,
    }
    return render(request, 'realizar_acordo.html', context)



@lojista_login_required
def quitar_parcela(request, titulo_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    titulo = get_object_or_404(Titulo, id=titulo_id)
    if request.method == 'POST':
        try:
            valor_recebido = float(request.POST.get('valorRecebido'))
            data_baixa = request.POST.get('dataBaixa')
            forma_pagamento = int(request.POST.get('formaPagamento'))

            # Atualizar o título
            titulo.valorRecebido = valor_recebido
            titulo.data_baixa = data_baixa
            titulo.forma_pag_Id = forma_pagamento  # Salvar a forma de pagamento no banco
            titulo.statusBaixa = 2  # Alterar status para Quitado
            titulo.save()

            #messages.success(request, f"Parcela {titulo.num_titulo} quitada com sucesso!")
        except Exception as e:
            messages.error(request, f"Erro ao quitar parcela: {e}")
    return redirect('listar_titulos_por_devedor', titulo.devedor_id)
    
    





def default_acordo(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    if request.user.is_authenticated:
        acordo = Acordo.objects.first()  # Substitua com sua lógica de seleção
        return {'acordo_id': acordo.id if acordo else None}
    return {}
    
@lojista_login_required
def gerar_pdf(request, titulo_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    """
    Gera um PDF detalhado para o título especificado usando ReportLab.
    """
    try:
        # Obtenha os dados do título e acordo
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    devedores.nome AS devedor_nome,
                    devedores.cpf,
                    devedores.cnpj,
                    titulo.valor AS valor_titulo,
                    titulo.juros,
                    core_empresa.nome_fantasia AS empresa_nome_fantasia,
                    core_empresa.cnpj AS empresa_cnpj,
                    core_acordo.valor_total_negociacao,
                    core_acordo.entrada,
                    core_acordo.qtde_prc,
                    core_acordo.data_entrada,
                    core_acordo.venc_primeira_parcela,
                    core_acordo.contato,
                    core_acordo.id AS acordo_id
                FROM 
                    devedores
                INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
                INNER JOIN titulo ON titulo.devedor_id = devedores.id
                INNER JOIN core_acordo ON core_acordo.titulo_id = titulo.id
                WHERE titulo.id = %s
            """, [titulo_id])
            acordo_data = cursor.fetchone()

        if not acordo_data:
            return HttpResponse("Acordo não encontrado.", status=404)

        # Mapear os dados para um dicionário
        acordo = {
            'devedor_nome': acordo_data[0],
            'cpf': acordo_data[1],
            'cnpj': acordo_data[2],
            'valor_titulo': acordo_data[3],
            'juros': acordo_data[4],
            'empresa_nome_fantasia': acordo_data[5],
            'empresa_cnpj': acordo_data[6],
            'valor_total_negociacao': acordo_data[7],
            'entrada': acordo_data[8],
            'qtde_prc': acordo_data[9],
            'data_entrada': acordo_data[10],
            'venc_primeira_parcela': acordo_data[11],
            'contato': acordo_data[12],
            'acordo_id': acordo_data[13],
        }

        # Obter as parcelas do acordo
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT parcela_numero, data_vencimento, valor
                FROM core_parcelamento
                WHERE acordo_id = %s
            """, [acordo['acordo_id']])
            parcelas = cursor.fetchall()

        # Criar um buffer para o PDF
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter

        # Adicionar título do PDF
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawCentredString(width / 2, height - 40, "ACORDO EXTRAJUDICIAL DE RENEGOCIAÇÃO DE DÍVIDA")

        # Adicionar detalhes do acordo
        pdf.setFont("Helvetica", 10)
       # pdf.drawString(50, height - 100, f"Nome do cliente: {acordo['devedor_nome']}")
       # pdf.drawString(50, height - 120, f"CPF/CNPJ: {acordo['cpf'] or acordo['cnpj']}")
       #  pdf.drawString(50, height - 140, f"Empresa: {acordo['empresa_nome_fantasia']}")
       # pdf.drawString(50, height - 160, f"CNPJ Empresa: {acordo['empresa_cnpj']}")
       # pdf.drawString(50, height - 180, f"Valor da Dívida: R$ {acordo['valor_titulo']:.2f}")
       # pdf.drawString(50, height - 200, f"Juros: R$ {acordo['juros']:.2f}")
       # Adicionar introdução ao contrato
       #pdf.drawString(50, height - 100, "ACORDO EXTRAJUDICIAL DE RENEGOCIAÇÃO DE DÍVIDA:")
        pdf.drawString(
            50, height - 70,
            f"Eu, {acordo['devedor_nome']}, portador do CPF/CNPJ {acordo['cpf'] or acordo['cnpj']}, confirmo a"
        )
        pdf.drawString(
            50, height - 90,
            f"Renegociação da dívida descrita acima em favor da empresa {acordo['empresa_nome_fantasia']},"
        )
        pdf.drawString(
            50, height - 110,
            f"De CNPJ {acordo['empresa_cnpj']}. Firmo este Contrato de Confissão e Renegociação de Dívida."
        )
        pdf.drawString(50, height - 130, f"Valor Total da Negociação: R$ {acordo['valor_total_negociacao']:.2f}")
        pdf.drawString(50, height - 150 , f"Entrada: R$ {acordo['entrada']:.2f}")
        pdf.drawString(50, height - 170, f"Quantidade de Parcelas: {acordo['qtde_prc']}")
        data_entrada_formatada = datetime.strptime(str(acordo['data_entrada']), '%Y-%m-%d').strftime('%d/%m/%Y')
        venc_primeira_parcela_formatada = datetime.strptime(str(acordo['venc_primeira_parcela']), '%Y-%m-%d').strftime('%d/%m/%Y')

        pdf.drawString(50, height - 190, f"Data da Entrada: {data_entrada_formatada}")
        pdf.drawString(50, height - 210, f"Vencimento da Primeira Parcela: {venc_primeira_parcela_formatada}")
       # pdf.drawString(50, height - 190, f"Data da Entrada: {acordo['data_entrada']}")
       # pdf.drawString(50, height - 210, f"Vencimento da Primeira Parcela: {acordo['venc_primeira_parcela']}")
        pdf.drawString(50, height - 230, f"Contato: {acordo['contato']}")

        
        

        # Adicionar tabela de parcelas
       # pdf.drawString(50, height - 440, "Parcelas:")
        pdf.line(50, height - 250, width - 50, height - 250)
        pdf.drawString(50, height - 270, "Parcela")
        pdf.drawString(150, height - 270, "Data de Vencimento")
        pdf.drawString(300, height - 270, "Valor")
        y = height - 290

        for parcela in parcelas:
            pdf.drawString(50, y, str(parcela[0]))
            pdf.drawString(150, y, parcela[1].strftime('%d/%m/%Y'))
            pdf.drawString(300, y, f"R$ {parcela[2]:.2f}")
            y -= 20

        # Assinatura
        pdf.drawString(50, y - 30, "Confirmo a renegociação nos termos acima.")
        pdf.line(70, y - 70, width - 70, y - 70)
        pdf.drawCentredString(width / 2, y - 80, f"{acordo['devedor_nome']}")
        pdf.drawCentredString(width / 2, y - 100, f"Assinatura")

        # Finalizar o PDF
        pdf.showPage()
        pdf.save()

        # Obter o conteúdo do PDF do buffer
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="acordo_titulo_{titulo_id}.pdf"'
        return response

    except Exception as e:
        # Log do erro para depuração
        print(f"Erro ao gerar PDF: {e}")
        return HttpResponse(f"Erro ao gerar PDF: {str(e)}", status=500)



@lojista_login_required
def listar_acordos(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    query = request.GET.get('q', '')

    sql_query = """
        SELECT 
            titulo.id AS titulo_id,
            titulo.valorRecebido,
            titulo.data_baixa,
            titulo.qtde_parcelas,
            titulo.total_acordo,
            titulo.dataVencimentoPrimeira,
            devedores.telefone1 AS contato,
            devedores.nome AS devedor_nome,
            core_empresa.nome_fantasia AS empresa_nome,
            devedores.cpf,
            devedores.cnpj,
            titulo.comprovante,
            titulo.contrato
        FROM 
            titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE 
            titulo.idTituloRef IS NULL 
            AND (titulo.statusBaixa = 2 OR titulo.statusBaixa = 3)
            AND devedores.empresa_id = %s
    """

    params = [empresa_id_sessao]
    if query:
        sql_query += """
            AND (
                devedores.nome LIKE %s OR
                core_empresa.nome_fantasia LIKE %s OR
                devedores.cpf LIKE %s OR
                devedores.cnpj LIKE %s
            )
        """
        params.extend([f'%{query}%'] * 4)

    sql_query += " ORDER BY titulo.id DESC"

    with connection.cursor() as cursor:
        cursor.execute(sql_query, params)
        rows = cursor.fetchall()

    acordos = [
        {
            'titulo_id': row[0],
            'valorRecebido': row[1],
            'data_baixa': row[2].strftime('%d/%m/%Y') if row[2] else '',
            'qtde_parcelas': row[3],
            'total_acordo': row[4],
            'dataVencimentoPrimeira': row[5].strftime('%d/%m/%Y') if row[5] else '',
            'contato': row[6],
            'devedor_nome': row[7],
            'empresa_nome': row[8],
            'cpf': row[9],
            'cnpj': row[10],
            'contrato': row[12],
            
        }
        for row in rows
    ]

    # Adiciona parcelas para cada título
    for acordo in acordos:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, valor, dataVencimento, data_baixa, statusBaixa, comprovante, contrato
                FROM titulo
                WHERE idTituloRef = %s
                ORDER BY dataVencimento ASC
            """, [acordo['titulo_id']])
            parcelas = cursor.fetchall()
            acordo['parcelas'] = [
                {
                    'id': parcela[0],
                    'valor': parcela[1],
                    'data_vencimento': parcela[2].strftime('%d/%m/%Y') if parcela[2] else '',
                    'data_baixa': parcela[3].strftime('%d/%m/%Y') if parcela[3] else '',
                    'status': 'Quitado' if parcela[4] == 2 else 'Pendente',
                    'status_baixa': parcela[4],
                    'comprovante': parcela[5],
                    'contrato': parcela[6],
                }
                for parcela in parcelas
            ]

            # Marcar parcelas sem IDs ou com dados inconsistentes
            for parcela in acordo['parcelas']:
                if not parcela['id']:
                    parcela['status'] = 'Invalida'

    paginator = Paginator(acordos, 10)  # Mostra 10 acordos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'acordos_listar.html', {'page_obj': page_obj, 'query': query})

from django.db import connection
from django.core.paginator import Paginator
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponseForbidden
from django.contrib import messages
from datetime import datetime
from decimal import Decimal, InvalidOperation

# Models usados
# from .models import Titulo  # já deve existir
# from .decorators import lojista_login_required  # já deve existir

# ----------------- Helpers -----------------
def _parse_valor_ptbr(s: str) -> Decimal:
    s = (s or '').strip()
    if not s:
        return Decimal('0')
    s = s.replace('.', '').replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal('0')

def _parse_date(s: str):
    s = (s or '').strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

# ----------------- Listagem -----------------
@lojista_login_required
def listar_acordos(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')
    query = request.GET.get('q', '')

    sql = """
      SELECT 
        t.id AS titulo_id,
        t.valorRecebido,
        t.data_baixa,
        t.qtde_parcelas,
        t.total_acordo,
        t.dataVencimentoPrimeira,
        d.telefone1 AS contato,
        d.nome AS devedor_nome,
        e.nome_fantasia AS empresa_nome,
        d.cpf, d.cnpj,
        t.comprovante, t.contrato
      FROM titulo t
      JOIN devedores d ON t.devedor_id = d.id
      JOIN core_empresa e ON d.empresa_id = e.id
      WHERE t.idTituloRef IS NULL
        AND (t.statusBaixa = 2 OR t.statusBaixa = 3)
        AND d.empresa_id = %s
    """
    params = [empresa_id_sessao]
    if query:
        sql += """
          AND (
            d.nome LIKE %s OR
            e.nome_fantasia LIKE %s OR
            d.cpf LIKE %s OR
            d.cnpj LIKE %s
          )
        """
        q = f'%{query}%'
        params.extend([q, q, q, q])

    sql += " ORDER BY t.id DESC"

    with connection.cursor() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()

    acordos = []
    for r in rows:
        acordo = {
            'titulo_id': r[0],
            'valorRecebido': r[1],
            'data_baixa': r[2].strftime('%d/%m/%Y') if r[2] else '',
            'qtde_parcelas': r[3],
            'total_acordo': r[4],
            'dataVencimentoPrimeira': r[5].strftime('%d/%m/%Y') if r[5] else '',
            'contato': r[6],
            'devedor_nome': r[7],
            'empresa_nome': r[8],
            'cpf': r[9],
            'cnpj': r[10],
            'contrato': r[12],
        }

        # Parcelas (filhas)
        with connection.cursor() as c2:
            c2.execute("""
              SELECT id, valor, dataVencimento, data_baixa, statusBaixa, comprovante, contrato
              FROM titulo
              WHERE idTituloRef = %s
              ORDER BY dataVencimento ASC, id ASC
            """, [acordo['titulo_id']])
            ps = c2.fetchall()

        parcelas = []
        for p in ps:
            dv = p[2]
            db = p[3]
            parcelas.append({
                'id': p[0],
                'valor': p[1],
                'data_vencimento_br': dv.strftime('%d/%m/%Y') if dv else '',
                'data_vencimento_iso': dv.strftime('%Y-%m-%d') if dv else '',
                'data_baixa_br': db.strftime('%d/%m/%Y') if db else '',
                'data_baixa_iso': db.strftime('%Y-%m-%d') if db else '',
                'status_baixa': p[4],     # 0/1/2/3
                'comprovante': p[5],
                'contrato': p[6],
            })
        acordo['parcelas'] = parcelas
        acordos.append(acordo)

    paginator = Paginator(acordos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'acordos_listar.html', {'page_obj': page_obj, 'query': query})

# ----------------- Editar parcela -----------------
@lojista_login_required
def parcela_editar(request, parcela_id: int):
    from .models import Titulo  # garante import local
    empresa_id = request.session.get('empresa_id_sessao')

    parcela = get_object_or_404(
        Titulo.objects.select_related('devedor'),
        id=parcela_id,
        devedor__empresa_id=empresa_id
    )
    if request.method != 'POST':
        return HttpResponseForbidden('Método não permitido')

    valor = _parse_valor_ptbr(request.POST.get('valor'))
    data_venc = _parse_date(request.POST.get('data_vencimento'))
    status_baixa = request.POST.get('status_baixa') or '0'
    data_baixa = _parse_date(request.POST.get('data_baixa'))

    try:
        status_baixa_int = int(status_baixa)
    except Exception:
        status_baixa_int = 0

    if status_baixa_int == 2 and not data_baixa:
        messages.error(request, 'Informe a data de pagamento para marcar como Quitado.')
        return redirect(request.POST.get('next') or 'acordos_listar')

    parcela.valor = valor
    parcela.dataVencimento = data_venc
    parcela.statusBaixa = status_baixa_int
    parcela.data_baixa = data_baixa if status_baixa_int == 2 else None
    parcela.save(update_fields=['valor','dataVencimento','statusBaixa','data_baixa'])

    messages.success(request, f'Parcela #{parcela.id} atualizada com sucesso!')
    return redirect(request.POST.get('next') or 'acordos_listar')

# ----------------- Excluir parcela -----------------
@lojista_login_required
def parcela_excluir(request, parcela_id: int):
    from .models import Titulo
    empresa_id = request.session.get('empresa_id_sessao')

    parcela = get_object_or_404(
        Titulo.objects.select_related('devedor'),
        id=parcela_id,
        devedor__empresa_id=empresa_id
    )
    if request.method != 'POST':
        return HttpResponseForbidden('Método não permitido')

    if parcela.statusBaixa == 2:
        messages.error(request, 'Parcela quitada não pode ser excluída.')
        return redirect(request.POST.get('next') or 'acordos_listar')

    if not parcela.idTituloRef:
        messages.error(request, 'Este registro não é uma parcela (é o acordo principal).')
        return redirect(request.POST.get('next') or 'acordos_listar')

    parcela.delete()
    messages.success(request, f'Parcela #{parcela_id} excluída com sucesso.')
    return redirect(request.POST.get('next') or 'acordos_listar')

def valor_por_extenso(valor):
    unidades = [
        '', 'um', 'dois', 'três', 'quatro', 'cinco', 'seis', 'sete', 'oito', 'nove'
    ]
    dezenas = [
        '', 'dez', 'vinte', 'trinta', 'quarenta', 'cinquenta', 'sessenta', 'setenta', 'oitenta', 'noventa'
    ]
    centenas = [
        '', 'cento', 'duzentos', 'trezentos', 'quatrocentos', 'quinhentos', 'seiscentos', 'setecentos', 'oitocentos', 'novecentos'
    ]
    especiais = {
        10: 'dez', 11: 'onze', 12: 'doze', 13: 'treze', 14: 'quatorze',
        15: 'quinze', 16: 'dezesseis', 17: 'dezessete', 18: 'dezoito', 19: 'dezenove'
    }

    def numero_por_extenso(n):
        if n == 0:
            return 'zero'
        elif n < 10:
            return unidades[n]
        elif n < 20:
            return especiais[n]
        elif n < 100:
            dezena, unidade = divmod(n, 10)
            return dezenas[dezena] + (f' e {unidades[unidade]}' if unidade else '')
        elif n < 1000:
            centena, resto = divmod(n, 100)
            if n == 100:
                return 'cem'
            return centenas[centena] + (f' e {numero_por_extenso(resto)}' if resto else '')
        else:
            milhar, resto = divmod(n, 1000)
            milhar_extenso = f'{numero_por_extenso(milhar)} mil' if milhar > 1 else 'mil'
            return milhar_extenso + (f' e {numero_por_extenso(resto)}' if resto else '')

    reais, centavos = divmod(round(valor * 100), 100)
    reais_extenso = f'{numero_por_extenso(reais)} real{"s" if reais > 1 else ""}' if reais else ''
    centavos_extenso = f'{numero_por_extenso(centavos)} centavo{"s" if centavos > 1 else ""}' if centavos else ''

    if reais and centavos:
        return f'{reais_extenso} e {centavos_extenso}'
    return reais_extenso or centavos_extenso



def gerar_contrato(request, titulo_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    # Obter o título principal
    titulo = get_object_or_404(Titulo, id=titulo_id, idTituloRef__isnull=True)

    # Obter o devedor e empresa associados
    devedor = titulo.devedor
    empresa = devedor.empresa

    # Obter as parcelas associadas (títulos filhos)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT id, valor, dataVencimento, data_baixa, nPrc
            FROM titulo
            WHERE idTituloRef = %s
            ORDER BY nPrc
        """, [titulo.id])
        parcelas = cursor.fetchall()

    # Função auxiliar para evitar erro com valores None
    def valor_extenso(valor):
        return valor_por_extenso(valor) if valor is not None else "Zero"

    # Preparar o contexto com os dados disponíveis
    context = {
        'devedores': {
            'nome': devedor.nome,
            'endereco': devedor.endereco,
            'cep': devedor.cep,
            'cidade': devedor.cidade,
            'uf': devedor.uf,
            'cpf': devedor.cpf,
        },
        'core_empresa': {
            'razao_social': empresa.razao_social,
            'endereco': empresa.endereco,
            'bairro': empresa.bairro,
            'cidade': empresa.cidade,
            'uf': empresa.uf,
            'cnpj': empresa.cnpj,
        },
        'titulo': {
            'valor_total_negociacao': titulo.total_acordo,
            'valor_total_negociacao_extenso': valor_extenso(titulo.total_acordo),
            'entrada': titulo.valorRecebido,
            'data_entrada': titulo.data_baixa,
            'entrada_extenso': valor_extenso(titulo.valorRecebido),
            'valor_por_parcela': titulo.parcelar_valor,
            'valor_por_parcela_extenso': valor_extenso(titulo.parcelar_valor),
            'qtde_prc': titulo.qtde_parcelas,
        },
        'parcelas': [
            {
                'parcela_numero': parcela[4],
                'data_vencimento_parcela': parcela[2],
                'valor': parcela[1],
                'valor_extenso': valor_extenso(parcela[1]),
            }
            for parcela in parcelas
        ],
    }

    # Renderizar o template
    return render(request, 'contrato_template.html', context)







@lojista_login_required
def realizar_baixa(request, titulo_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    print(f"Requisição recebida: {request.method}")
    titulo = get_object_or_404(Titulo, id=titulo_id)

    forma_pagamento_map = {
        0: "Pix",
        1: "Dinheiro",
        2: "Cartão de Débito",
        3: "Cartão de Crédito",
        4: "Cheque",
        5: "Depósito em Conta",
        6: "Pagamento na Loja",
        7: "Boleto Bancário",
        8: "Duplicata",
    }

    if request.method == 'POST':
        print("Received POST data:", request.POST)  # Debug incoming data

        try:
            tipo_baixa = request.POST.get('tipo_baixa')
            forma_pagamento_key = int(request.POST.get('forma_pagamento', 0))
            forma_pagamento = forma_pagamento_map.get(forma_pagamento_key, "Indefinido")

            print(f"Tipo de Baixa: {tipo_baixa}, Forma de Pagamento: {forma_pagamento}")  # Debug the type and payment method

            if tipo_baixa == 'Quitação':
                valor_quitacao = float(request.POST.get('valor_quitacao', 0))
                data_pagamento = request.POST.get('data_pagamento')

                print(f"Quitação: valor={valor_quitacao}, data_pagamento={data_pagamento}")

                # Atualiza os valores no modelo Titulo para Quitação
                Titulo.objects.filter(id=titulo.id).update(
                    data_baixa=data_pagamento,
                    valorRecebido=valor_quitacao,
                    forma_pag_Id=forma_pagamento_key,
                    statusBaixa=2  # Quitado
                )

                titulo.data_baixa = data_pagamento
                titulo.valorRecebido = valor_quitacao
                titulo.forma_pag_Id = forma_pagamento_key
                titulo.statusBaixa = 2
                titulo.save()

                print(f"Título atualizado com sucesso: {titulo}")
                #messages.success(request, f"Baixa realizada com sucesso para Quitação via {forma_pagamento}!")

            elif tipo_baixa == 'Parcela':
                valor_parcela = float(request.POST.get('valor_parcela', 0))
                data_pagamento = request.POST.get('data_pagamento')

                print(f"Parcela: valor={valor_parcela}, data_pagamento={data_pagamento}")

                # Incrementa o valor recebido e atualiza o status do título
                Titulo.objects.filter(id=titulo.id).update(
                    valorRecebido=(titulo.valorRecebido or 0) + valor_parcela,
                    data_baixa=data_pagamento,
                    forma_pag_Id=forma_pagamento_key,
                    statusBaixa=1 if (titulo.valorRecebido or 0) + valor_parcela < titulo.valor else 2  # Parcial ou Quitado
                )

                titulo.valorRecebido = (titulo.valorRecebido or 0) + valor_parcela
                titulo.data_baixa = data_pagamento
                titulo.forma_pag_Id = forma_pagamento_key
                titulo.statusBaixa = 1 if titulo.valorRecebido < titulo.valor else 2
                titulo.save()

                print(f"Título atualizado para pagamento de parcela: {titulo}")
                #messages.success(request, f"Pagamento de parcela registrado com sucesso via {forma_pagamento}!")

            else:
                print("Tipo de Baixa inválido.")
                messages.error(request, "Tipo de Baixa inválido.")

            return redirect('titulos_listar')

        except Exception as e:
            print(f"Erro ao salvar Baixa: {e}")
            messages.error(request, f"Erro ao salvar Baixa: {e}")
            return redirect('realizar_baixa', titulo_id=titulo_id)

    diferenca_dias = (datetime.today().date() - titulo.dataVencimento).days
    juros_totais = (titulo.valor * 0.08 * (diferenca_dias / 30)) if diferenca_dias > 0 else 0

    context = {
        'titulo': titulo,
        'juros_totais': juros_totais,
        'data_vencimento_formatada': titulo.dataVencimento.strftime('%d/%m/%Y'),
    }
    return render(request, 'realizar_baixa.html', context)


    
    
    
@lojista_login_required    
def listar_parcelamentos(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    # Obter o termo de pesquisa
    query = request.GET.get('q', '')

    # Construir a consulta SQL com filtro, se aplicável
    sql_query = """
        SELECT 
            core_parcelamento.id, 
            core_parcelamento.parcela_numero, 
            core_parcelamento.valor,
            core_parcelamento.data_vencimento,
            core_parcelamento.data_vencimento_parcela,
            core_parcelamento.status, 
            core_parcelamento.created_at,
            core_parcelamento.acordo_id, 
            core_acordo.entrada, 
            core_acordo.qtde_prc, 
            core_acordo.contato, 
            core_acordo.titulo_id,
            devedores.nome AS devedor_nome,
            core_empresa.nome_fantasia AS empresa_nome_fantasia,
            core_parcelamento.forma_pagamento,
            devedores.cpf,
            devedores.cnpj
        FROM 
            core_parcelamento
        INNER JOIN 
            core_acordo 
        ON 
            core_parcelamento.acordo_id = core_acordo.id
        INNER JOIN 
            titulo
        ON 
            core_acordo.titulo_id = titulo.id
        INNER JOIN 
            devedores
        ON 
            titulo.devedor_id = devedores.id
        INNER JOIN 
            core_empresa
        ON 
            devedores.empresa_id = core_empresa.id
        WHERE 1=1
    """

    # Adicionar filtro baseado no termo de pesquisa
    params = []
    if query:
        sql_query += """
            AND (
                core_parcelamento.parcela_numero LIKE %s OR
                devedores.nome LIKE %s OR
                core_empresa.nome_fantasia LIKE %s OR
                core_acordo.contato LIKE %s OR
                core_acordo.titulo_id LIKE %s
            )
        """
        params = [f'%{query}%'] * 5

    sql_query += " ORDER BY core_parcelamento.parcela_numero ASC"

    # Executar a consulta
    with connection.cursor() as cursor:
        cursor.execute(sql_query, params)
        rows = cursor.fetchall()

    # Mapear os resultados para uma estrutura legível no template
    parcelamentos = [
        {
            "id": row[0],
            "parcela_numero": f"{row[1]}/{row[9]}",  # Formata parcela_numero / qtde_prc
            "valor": row[2],
            "data_vencimento": row[3],
            "data_vencimento_parcela": row[4],
            "status": row[5],
            "created_at": row[6],
            "acordo_id": row[7],
            "entrada": row[8],
            "qtde_prc": row[9],
            "contato": row[10],
            "titulo_id": row[11],
            "devedor_nome": row[12],
            "empresa_nome_fantasia": row[13],
            "agendamento_forma_pagamento": row[14],
            "cpf": row[15],
            "cnpj": row[16],
        }
        for row in rows
    ]

    # Paginação (20 itens por página)
    paginator = Paginator(parcelamentos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Renderizar o template com os dados
    return render(
        request,
        'parcelamentos_listar.html',  # Nome do template para exibir os parcelamentos
        {'page_obj': page_obj, 'query': query}
    )


# ========================= Relatório de Honorários =========================
from django.db.models import Q, F, Sum, Value as V
from django.db.models.functions import Coalesce
from decimal import Decimal
import csv
from urllib.parse import urlencode
from django.urls import reverse
from django.http import FileResponse, Http404
import mimetypes
import logging

logger = logging.getLogger(__name__)


def calcular_comissao_por_max_atraso_historico(devedor_id):
    """
    Calcula o percentual de comissão baseado no MAIOR ATRASO HISTÓRICO do devedor.
    Retorna o percentual como Decimal (ex: 0.21 para 21%).
    
    Lógica:
    - Busca TODOS os títulos já quitados do devedor
    - Calcula o maior número de dias de atraso histórico
    - Retorna o percentual baseado na faixa:
      - 30-90 dias: 9%
      - 91-180 dias: 15%
      - 181-720 dias: 21%
      - 721-1825 dias: 30%
      - >= 1826 dias: 40%
      - < 30 dias: 0%
    """
    from django.db import connection
    
    try:
        with connection.cursor() as cursor:
            # CTE para calcular o maior atraso histórico do devedor
            cursor.execute("""
                WITH max_atraso AS (
                    SELECT
                        d.id AS devedor_id,
                        MAX(GREATEST(
                            0,
                            DATEDIFF(
                                t2.data_baixa,
                                COALESCE(t2.dataVencimentoReal, t2.dataVencimento, t2.dataVencimentoPrimeira)
                            )
                        )) AS max_dias
                    FROM titulo t2
                    JOIN devedores d ON d.id = t2.devedor_id
                    JOIN core_empresa e ON e.id = d.empresa_id
                    WHERE t2.data_baixa IS NOT NULL
                      AND e.status_empresa = 1
                      AND d.id = %s
                    GROUP BY d.id
                )
                SELECT max_dias
                FROM max_atraso
            """, [devedor_id])
            
            result = cursor.fetchone()
            
            if not result or result[0] is None:
                return Decimal('0.00')  # Sem histórico, retorna 0%
            
            max_dias = int(result[0])
            
            # Determinar percentual baseado na faixa
            if max_dias < 30:
                return Decimal('0.00')  # 0%
            elif max_dias <= 90:
                return Decimal('0.09')  # 9%
            elif max_dias <= 180:
                return Decimal('0.15')  # 15%
            elif max_dias <= 720:
                return Decimal('0.21')  # 21%
            elif max_dias <= 1825:
                return Decimal('0.30')  # 30%
            else:
                return Decimal('0.40')  # 40%
                
    except Exception as e:
        logger.error(f"Erro ao calcular comissão por max atraso histórico para devedor {devedor_id}: {e}")
        return Decimal('0.00')


@lojista_login_required
def relatorio_honorarios(request):
    """
    Relatório de Honorários (Quitados) por período com filtros.
    - Filtra sempre pela empresa do lojista via empresa_id na sessão.
    - Filtros: data início, data fim, consultor (empresa.operador), credor (empresa), devedor (nome), cpf/cnpj.
    - Totais: Quitado, Comissão (15%), Operador (25% da comissão), Líquido.
    """
    empresa_id = request.session.get('empresa_id_sessao')
    if not empresa_id:
        return redirect('login')

    # Parâmetros
    dt_ini_str = request.GET.get('data_inicio', '')
    dt_fim_str = request.GET.get('data_fim', '')
    consultor = (request.GET.get('consultor') or '').strip()
    credor_id = request.GET.get('credor')  # id da empresa (opcional)
    devedor_nome = (request.GET.get('devedor') or '').strip()
    cpf_cnpj = (request.GET.get('cpf_cnpj') or '').strip()

    # Datas padrão: últimos 7 dias
    from datetime import timedelta
    hojed = now().date()
    default_ini = hojed - timedelta(days=7)

    def _parse(s):
        from datetime import datetime as _dt
        s = (s or '').strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return _dt.strptime(s, fmt).date()
            except Exception:
                pass
        return None

    dt_ini = _parse(dt_ini_str) or default_ini
    dt_fim = _parse(dt_fim_str) or hojed

    # Query base: títulos quitados (statusBaixa=2) da empresa do lojista
    # Usamos Titulo.join Devedor, Empresa para filtrar e exibir
    # IMPORTANTE: Incluir 'devedor__empresa__plano' no select_related para evitar queries extras
    titulos = (
        Titulo.objects
        .select_related('devedor', 'devedor__empresa', 'devedor__empresa__plano', 'empresa', 'empresa__plano')
        .filter(
            devedor__empresa_id=empresa_id,
            statusBaixa=2,
            data_baixa__isnull=False,
            data_baixa__range=(dt_ini, dt_fim),
            valorRecebido__isnull=False
        )
    )

    if consultor:
        # consultor ligado ao operador da empresa do devedor
        titulos = titulos.filter(devedor__empresa__operador__icontains=consultor)
    if credor_id and credor_id.isdigit():
        # credor pela empresa do devedor (titulo.empresa pode ser nulo)
        titulos = titulos.filter(devedor__empresa_id=int(credor_id))
    if devedor_nome:
        titulos = titulos.filter(devedor__nome__icontains=devedor_nome)
    if cpf_cnpj:
        titulos = titulos.filter(Q(devedor__cpf__icontains=cpf_cnpj) | Q(devedor__cnpj__icontains=cpf_cnpj))

    # Montagem de linhas
    linhas = []
    total_quitado = Decimal('0')
    total_comissao = Decimal('0')
    total_liquido = Decimal('0')
    forma_map = {
        0: "Pix",
        1: "Dinheiro",
        2: "Cartão de Débito",
        3: "Cartão de Crédito",
        4: "Cheque",
        5: "Depósito em Conta",
        6: "Pagamento na Loja",
        7: "Boleto Bancário",
        8: "Duplicata",
    }
    for t in titulos:
        pago = Decimal(str(t.valorRecebido or 0))
        # Calcular comissão baseada no MAIOR ATRASO HISTÓRICO do devedor
        comissao_percent = calcular_comissao_por_max_atraso_historico(t.devedor_id)
        honor = (pago * comissao_percent).quantize(Decimal('0.01'))
        liquido = (pago - honor).quantize(Decimal('0.01'))
        principal = Decimal(str(t.valor or 0)).quantize(Decimal('0.01'))
        total_quitado += pago
        total_comissao += honor
        total_liquido += liquido
        doc = t.devedor.cpf or t.devedor.cnpj or ''
        empresa_obj = getattr(t.devedor, 'empresa', None) or t.empresa
        consultor_nome = getattr(empresa_obj, 'operador', '') if empresa_obj else ''
        credor_disp = empresa_obj.nome_fantasia if empresa_obj else ''
        venc_br = t.dataVencimento.strftime('%d/%m/%Y') if t.dataVencimento else ''
        pagto_br = t.data_baixa.strftime('%d/%m/%Y') if t.data_baixa else ''
        forma_str = forma_map.get(t.forma_pag_Id, '') if t.forma_pag_Id is not None else ''
        parc_str = f"{t.nPrc or 1}/{t.qtde_parcelas or 1}"
        comprovante_url = ''
        try:
            if t.comprovante and getattr(t.comprovante, 'name', ''):
                comprovante_url = reverse('relatorio_honorarios_comprovante', args=[t.id])
        except Exception:
            comprovante_url = ''
        linhas.append({
            'titulo_id': t.id,
            'devedor': t.devedor.nome or t.devedor.razao_social or f'#{t.devedor_id}',
            'documento': doc,
            'credor_display': credor_disp,
            'consultor': consultor_nome,
            'venc_br': venc_br,
            'pagto_br': pagto_br,
            'forma': forma_str,
            'parcelas': parc_str,
            'principal': principal,
            'pago': pago,
            'honorarios': honor,
            'liquido': liquido,
            # Strings formatadas para uso direto no modal (pt-BR)
            'principal_str': f"{principal:.2f}".replace('.', ','),
            'pago_str': f"{pago:.2f}".replace('.', ','),
            'honorarios_str': f"{honor:.2f}".replace('.', ','),
            'liquido_str': f"{liquido:.2f}".replace('.', ','),
            'comprovante_url': comprovante_url,
        })

    # Totais já foram acumulados no loop acima
    # Garantir arredondamento final para evitar problemas de precisão
    total_comissao = total_comissao.quantize(Decimal('0.01'))
    total_liquido = total_liquido.quantize(Decimal('0.01'))

    # Paginação
    paginator = Paginator(linhas, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Lista de credores (somente da empresa do lojista = próprio id)
    # Aqui exibimos apenas a própria empresa (lojista vê só dele)
    from .models import Empresa
    credores = Empresa.objects.filter(id=empresa_id)

    # Querystring para export (preserva filtros)
    qs_export = request.GET.copy()
    if 'page' in qs_export:
        qs_export.pop('page')
    export_url = f"{reverse('relatorio_honorarios_exportar')}?{qs_export.urlencode()}"

    context = {
        'page_obj': page_obj,
        'data_inicio': dt_ini.strftime('%Y-%m-%d'),
        'data_fim': dt_fim.strftime('%Y-%m-%d'),
        'consultor': consultor,
        'credor_id': int(credor_id) if (credor_id and credor_id.isdigit()) else None,
        'devedor': devedor_nome,
        'cpf_cnpj': cpf_cnpj,
        'credores': credores,
        'total_quitado': total_quitado,
        'total_comissao': total_comissao,
        'total_liquido': total_liquido,
        'export_url': export_url,
    }
    return render(request, 'relatorio_honorarios.html', context)


@lojista_login_required
def relatorio_honorarios_exportar(request):
    """Exporta o relatório em PDF respeitando os filtros."""
    empresa_id = request.session.get('empresa_id_sessao')
    if not empresa_id:
        return redirect('login')

    # Mesmos filtros da view principal
    dt_ini = request.GET.get('data_inicio')
    dt_fim = request.GET.get('data_fim')
    consultor = (request.GET.get('consultor') or '').strip()
    credor_id = request.GET.get('credor')
    devedor_nome = (request.GET.get('devedor') or '').strip()
    cpf_cnpj = (request.GET.get('cpf_cnpj') or '').strip()

    def _parse(s):
        from datetime import datetime as _dt
        s = (s or '').strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return _dt.strptime(s, fmt).date()
            except Exception:
                pass
        return None

    from datetime import timedelta
    hojed = now().date()
    default_ini = hojed - timedelta(days=7)
    d_ini = _parse(dt_ini) or default_ini
    d_fim = _parse(dt_fim) or hojed

    titulos = (
        Titulo.objects
        .select_related('devedor', 'devedor__empresa', 'devedor__empresa__plano', 'empresa', 'empresa__plano')
        .filter(
            devedor__empresa_id=empresa_id,
            statusBaixa=2,
            data_baixa__isnull=False,
            data_baixa__range=(d_ini, d_fim),
            valorRecebido__isnull=False
        )
    )
    if consultor:
        titulos = titulos.filter(devedor__empresa__operador__icontains=consultor)
    if credor_id and credor_id.isdigit():
        titulos = titulos.filter(devedor__empresa_id=int(credor_id))
    if devedor_nome:
        titulos = titulos.filter(devedor__nome__icontains=devedor_nome)
    if cpf_cnpj:
        titulos = titulos.filter(Q(devedor__cpf__icontains=cpf_cnpj) | Q(devedor__cnpj__icontains=cpf_cnpj))

    # Preparar dados
    linhas_dados = []
    total_quitado = Decimal('0')
    total_comissao = Decimal('0')
    total_liquido = Decimal('0')

    for t in titulos:
        pago = Decimal(str(t.valorRecebido or 0))
        # Calcular comissão baseada no MAIOR ATRASO HISTÓRICO do devedor
        comissao_percent = calcular_comissao_por_max_atraso_historico(t.devedor_id)
        honor = (pago * comissao_percent).quantize(Decimal('0.01'))
        liquido = (pago - honor).quantize(Decimal('0.01'))
        empresa_obj = getattr(t.devedor, 'empresa', None) or t.empresa
        credor_display = f"{empresa_obj.id} - {empresa_obj.nome_fantasia}" if empresa_obj else ''
        devedor_nome_display = t.devedor.nome or t.devedor.razao_social or f'#{t.devedor_id}'
        
        # Formatar valores no padrão brasileiro (R$ 1.234,56)
        pago_str = f"R$ {pago:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        honor_str = f"R$ {honor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        liquido_str = f"R$ {liquido:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        linhas_dados.append([
            devedor_nome_display,
            credor_display,
            pago_str,
            honor_str,
            liquido_str,
        ])
        total_quitado += pago
        total_comissao += honor
        total_liquido += liquido

    # Criar PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, 
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=30)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#212529'),
        spaceAfter=20,
        alignment=1,  # Center
    )
    
    # Conteúdo
    story = []
    
    # Título
    story.append(Paragraph("Relatório de Honorários", title_style))
    
    # Informações do período
    periodo_text = f"Período: {d_ini.strftime('%d/%m/%Y')} a {d_fim.strftime('%d/%m/%Y')}"
    story.append(Paragraph(periodo_text, styles['Normal']))
    story.append(Paragraph("<br/>", styles['Normal']))
    
    # Cabeçalho da tabela
    header = [['Devedor', 'Credor', 'Pago', 'Honorários', 'Líquido']]
    
    # Estilo para cabeçalho
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Helvetica-Bold',
        textColor=colors.whitesmoke,
    )
    
    # Dados - converter para Paragraph para melhor formatação
    data = []
    for row in header:
        data.append([Paragraph(cell, header_style) for cell in row])
    
    for row in linhas_dados:
        data.append([Paragraph(cell, styles['Normal']) for cell in row])
    
    # Linha de totais - formatar valores no padrão brasileiro
    total_quitado_str = f"R$ {total_quitado:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    total_comissao_str = f"R$ {total_comissao:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    total_liquido_str = f"R$ {total_liquido:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    
    data.append([
        Paragraph('<b>TOTAIS</b>', styles['Normal']),
        Paragraph('', styles['Normal']),
        Paragraph(f"<b>{total_quitado_str}</b>", styles['Normal']),
        Paragraph(f"<b>{total_comissao_str}</b>", styles['Normal']),
        Paragraph(f"<b>{total_liquido_str}</b>", styles['Normal']),
    ])
    
    # Criar tabela - ajustar larguras das colunas (largura útil ~490 pontos)
    table = Table(data, colWidths=[140, 140, 90, 90, 90])
    
    # Estilo da tabela
    table.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#495057')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        # Linhas de dados
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -2), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 8),
        # Linha de totais
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
    ]))
    
    story.append(table)
    story.append(Paragraph("<br/>", styles['Normal']))
    
    # Informações adicionais
    info_text = f"Comissão: {int(COMISSAO_PERCENT * 100)}% | Total de registros: {len(linhas_dados)}"
    story.append(Paragraph(info_text, styles['Normal']))
    
    # Construir PDF
    doc.build(story)
    
    # Preparar resposta
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_honorarios.pdf"'

    return response

@lojista_login_required
def relatorio_honorarios_comprovante(request, titulo_id: int):
    """Entrega o comprovante associado ao título garantindo o vínculo com a empresa logada."""
    empresa_id = request.session.get('empresa_id_sessao')
    titulo = get_object_or_404(
        Titulo.objects.select_related('devedor'),
        id=titulo_id,
        devedor__empresa_id=empresa_id
    )
    comprovante = titulo.comprovante
    if not comprovante or not getattr(comprovante, 'name', ''):
        raise Http404("Comprovante não disponível.")

    candidate_paths = []
    try:
        candidate_paths.append(comprovante.path)
    except Exception:
        pass

    if settings.MEDIA_ROOT:
        candidate_paths.append(os.path.join(settings.MEDIA_ROOT, comprovante.name))

    alt_media_root = getattr(settings, 'ALT_MEDIA_ROOT', '/home/app_operador/media')
    if alt_media_root:
        candidate_paths.append(os.path.join(alt_media_root, comprovante.name))

    file_path = next((p for p in candidate_paths if p and os.path.exists(p)), None)
    if not file_path:
        raise Http404("Comprovante não encontrado no servidor.")

    content_type, _ = mimetypes.guess_type(file_path)
    response = FileResponse(open(file_path, 'rb'), content_type=content_type or 'application/octet-stream')
    filename = os.path.basename(file_path)
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response

# core/views.py
from django.shortcuts import render
from core.decorators import lojista_login_required

@lojista_login_required
def followups_listar(request):
    return render(request, "follow_ups_listar.html", {"page_title": "Follow-ups"})



@lojista_login_required
def gerar_recibo(request, titulo_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    # Buscar o título pelo ID e verificar se está quitado
    titulo = get_object_or_404(Titulo, id=titulo_id, statusBaixa=2)

    # Obter o devedor associado ao título
    devedor = get_object_or_404(Devedor, id=titulo.devedor_id)

    # Obter a empresa associada ao devedor
    empresa = get_object_or_404(Empresa, id=devedor.empresa_id)

    # Mapear IDs de forma de pagamento para nomes legíveis
    forma_pagamento_map = {
        0: "Pix",
        1: "Dinheiro",
        2: "Cartão de Débito",
        3: "Cartão de Crédito",
        4: "Cheque",
        5: "Depósito em Conta",
        6: "Pagamento na Loja",
        7: "Boleto Bancário",
        8: "Duplicata",
    }

    # Obter a forma de pagamento legível
    forma_pagamento_legivel = forma_pagamento_map.get(titulo.forma_pag_Id, "Não definido")

    # Preparar o contexto para o recibo
    context = {
        'empresa': {
            'nome_fantasia': empresa.nome_fantasia or "N/A",
            'razao_social': empresa.razao_social or "N/A",
            'cnpj': empresa.cnpj or "N/A",
            'endereco': f"{empresa.endereco}, {empresa.numero}" if empresa.endereco and empresa.numero else "N/A",
            'cidade': empresa.cidade or "N/A",
            'uf': empresa.uf or "N/A",
            'email': empresa.email or "N/A",
            'banco': empresa.banco or "N/A",
        },
        'devedor': {
            'nome': devedor.nome or devedor.razao_social,
            'cpf_cnpj': devedor.cpf if devedor.cpf else devedor.cnpj,
        },
        'parcela': {
            'numero': titulo.nPrc,
            'qtde_total': titulo.qtde_parcelas,
            'data_vencimento': titulo.dataVencimento,
            'data_pagamento': titulo.data_baixa,
            'valor_pago': titulo.valorRecebido,
            'forma_pagamento': forma_pagamento_legivel,  # Nome legível da forma de pagamento
        },
        'autenticacao': f"Recibo gerado em {titulo.data_baixa.strftime('%d/%m/%Y')}" if titulo.data_baixa else "N/A",
    }

    # Renderizar o recibo
    return render(request, 'recibo.html', context)


@lojista_login_required
@require_POST
def pagar_parcela(request, parcela_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    # Obter os dados do formulário
    valor_pago = float(request.POST.get('valor_pago', 0))
    data_baixa = request.POST.get('data_baixa')
    forma_pagamento = request.POST.get('forma_pagamento')  # Captura a forma de pagamento

    # Validar se a parcela existe no banco de dados
    with connection.cursor() as cursor:
        cursor.execute("SELECT valor FROM core_parcelamento WHERE id = %s", [parcela_id])
        parcela = cursor.fetchone()

        if not parcela:
            messages.error(request, "Parcela não encontrada.")
            return redirect('listar_parcelamentos')

        valor_original = parcela[0]

        # Atualizar a parcela com os dados fornecidos
        cursor.execute("""
            UPDATE core_parcelamento
            SET status = %s, data_baixa = %s, forma_pagamento = %s
            WHERE id = %s
        """, ["Quitado", data_baixa, forma_pagamento, parcela_id])

    # Exibir mensagem de sucesso e redirecionar
    #messages.success(request, f"Parcela {parcela_id} atualizada com sucesso.")
    return redirect('listar_parcelamentos')



@lojista_login_required
def listar_agendamentos(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão

    # Verifica se o empresa_id_sessao está presente na sessão
    if not empresa_id_sessao:
        messages.error(request, "Sessão expirada ou inválida. Por favor, faça login novamente.")
        return redirect('login')  # Substitua 'login' pelo nome correto da sua URL de login

    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()  # Filtro para status
    page_number = request.GET.get('page', 1)  # Número da página atual

    # Obter todos os agendamentos relacionados à empresa do lojista logado
    agendamentos = Agendamento.objects.select_related('devedor', 'empresa').filter(
        empresa__id=empresa_id_sessao
    )

    # Aplicar filtro de busca, se houver
    if query:
        agendamentos = agendamentos.filter(
            Q(devedor__nome__icontains=query) |
            Q(devedor__cpf__icontains=query) |
            Q(devedor__cnpj__icontains=query)
        )

    # Aplicar filtro de status, se houver
    if status_filter:
        agendamentos = agendamentos.filter(status=status_filter)

    # Ordenar por status, colocando "Pendente" primeiro
    agendamentos = agendamentos.annotate(
        status_priority=Case(
            When(status='Pendente', then=1),
            default=2,
            output_field=IntegerField()
        )
    ).order_by(
        'status_priority',
        'data_retorno'  # Depois ordena por data de retorno
    )

    # Paginação: 10 itens por página
    paginator = Paginator(agendamentos, 10)
    agendamentos_paginados = paginator.get_page(page_number)

    # Renderizar o template
    return render(request, 'agendamentos_listar.html', {
        'agendamentos': agendamentos_paginados,  # Agendamentos paginados
        'query': query,
        'status_filter': status_filter,  # Enviar o filtro para o template
    })


@lojista_login_required    
def finalizar_agendamento(request, agendamento_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    if request.method == 'POST':
        agendamento.status = 'Finalizado'
        agendamento.save()
       # messages.success(request, 'Agendamento finalizado com sucesso.')
    return redirect('listar_agendamentos')



@lojista_login_required
def criar_agendamento(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    """
    View para criar um novo agendamento.
    """
    # Recupera os devedores para o template com select explícito
    devedores = Devedor.objects.select_related('devedores').filter(empresa__id=empresa_id_sessao).values(
        'id',
        'nome',
        'empresa__id',
        'empresa__nome_fantasia',
        'telefone'
    )

    devedores_com_empresas = [
        {
            "id": devedor['id'],
            "nome": devedor['nome'],
            "empresa_id": devedor['empresa__id'],
            "nome_fantasia": devedor['empresa__nome_fantasia'],
            "telefone": devedor['telefone']
        }
        for devedor in devedores
    ]

    # Recupera o nome da empresa da sessão
    empresa_nome_fantasia = ""
    if empresa_id_sessao:
        empresa = Empresa.objects.filter(id=empresa_id_sessao).values('nome_fantasia').first()
        if empresa:
            empresa_nome_fantasia = empresa['nome_fantasia']

    if request.method == 'POST':
        try:
            # Captura os dados do formulário
            devedor_id = request.POST['devedor_id']
            telefone = request.POST['telefone']
            data_abertura = make_aware(datetime.strptime(request.POST['data_abertura'], "%Y-%m-%dT%H:%M"))
            data_retorno = make_aware(datetime.strptime(request.POST['data_retorno'], "%Y-%m-%dT%H:%M"))
            operador = request.POST.get('operador', '')
            assunto = request.POST.get('assunto', '')  # Captura o assunto

            # Recupera o devedor e a empresa associada
            devedor = get_object_or_404(Devedor, id=devedor_id)
            empresa = devedor.empresa

            # Cria o agendamento
            Agendamento.objects.create(
                devedor=devedor,
                empresa=empresa,
                telefone=telefone,
                data_abertura=data_abertura,
                data_retorno=data_retorno,
                operador=operador,
                assunto=assunto,  # Atribui o assunto aqui
                status='Pendente',  # Status inicial
            )

            # Exibe mensagem de sucesso
            # messages.success(request, "Agendamento criado com sucesso!")
            print("Agendamento criado com sucesso!")  # Log para depuração

            # Redireciona para a lista de agendamentos
            return redirect('listar_agendamentos')

        except Exception as e:
            # Em caso de erro
            messages.error(request, f"Erro ao criar agendamento: {e}")
            print(f"Erro ao criar agendamento: {e}")  # Log para depuração

    return render(request, 'agendamentos_criar.html', {
        'devedores': devedores_com_empresas,
        'empresa_id': empresa_id_sessao,
        'nome_fantasia': empresa_nome_fantasia
    })





@lojista_login_required
def anexar_comprovante(request, parcela_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    # Obtenha a parcela diretamente da tabela Titulo
    parcela = get_object_or_404(Titulo, id=parcela_id, idTituloRef__isnull=False)

    if request.method == 'POST' and 'comprovante' in request.FILES:
        comprovante = request.FILES['comprovante']

        # Gera um nome de arquivo único com UUID e preserva a extensão original
        extension = os.path.splitext(comprovante.name)[1]  # Obtém a extensão do arquivo
        unique_filename = f"{uuid.uuid4()}{extension}"     # Nome único com extensão original

        # Salva o arquivo com o nome único no campo 'comprovante' da parcela
        parcela.comprovante.save(unique_filename, comprovante)
        parcela.save()  # Salva as mudanças no banco de dados

        messages.success(request, "Comprovante anexado com sucesso!")
        return redirect('acordos_listar')  # Redireciona de volta à lista de parcelamentos

    messages.error(request, "Falha ao anexar o comprovante. Tente novamente.")
    return redirect('acordos_listar')



def baixar_comprovante(request, titulo_id):
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    titulo = Titulo.objects.get(id=titulo_id)
    if titulo.comprovante:
        comprovante_path = titulo.comprovante.path  # Usando o atributo .path que já considera o MEDIA_ROOT

        with open(comprovante_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(comprovante_path)}"'
            return response
    else:
        return HttpResponse("Nenhum comprovante disponível.", status=404)



@lojista_login_required
def detalhes_devedor(request, titulo_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    titulo = get_object_or_404(Titulo, id=titulo_id)
    # Verifica se o título pertence à empresa na sessão
    if titulo.empresa_id != empresa_id_sessao:
        return HttpResponseForbidden("Você não tem permissão para acessar detalhes de títulos de outra empresa.")
    devedor = titulo.devedor
    hoje = now().date()

    if not devedor:
        messages.error(request, "Devedor associado a este título não foi encontrado.")
        return redirect('lista_titulos')

    # Separate QuerySets for 'entradas' and 'títulos associados'
    titulos_entrada = Titulo.objects.filter(devedor=devedor, idTituloRef__isnull=True).select_related('empresa', 'devedor')
    titulos_associados = Titulo.objects.filter(devedor=devedor, idTituloRef__isnull=False).select_related('empresa', 'devedor')

    # Calculate dues and apply filters on all titles associated with the debtor
    titulos = Titulo.objects.filter(devedor=devedor).select_related('empresa', 'devedor')
    for titulo in titulos:
        if titulo.dataVencimento and titulo.dataVencimento < hoje:
            diferenca_dias = (hoje - titulo.dataVencimento).days
            juros_mensais = 0.08
            juros_totais = (titulo.valor * juros_mensais) * (diferenca_dias / 30)
            titulo.juros = round(juros_totais, 2)
            titulo.dias_atraso = diferenca_dias
            titulo.save(update_fields=['juros', 'dias_atraso'])

    # Aggregate totals for display
    total_quitado = titulos.filter(statusBaixa=2).aggregate(total=Sum('valor'))['total'] or 0
    total_negociado = titulos.filter(statusBaixa=3).aggregate(total=Sum('valor'))['total'] or 0
    total_pendente = titulos.filter(statusBaixa=0).aggregate(total=Sum('valor'))['total'] or 0

    # Payment method mapping
    forma_pagamento_map = {
        0: "Pix",
        1: "Dinheiro",
        2: "Cartão de Débito",
        3: "Cartão de Crédito",
        4: "Cheque",
        5: "Depósito em Conta",
        6: "Pagamento na Loja",
        7: "Boleto Bancário",
        8: "Duplicata",
    }

    # Additional context: company, schedules, follow-ups
    empresa = getattr(devedor, 'empresa', None)
    agendamentos = Agendamento.objects.filter(devedor=devedor)
    follow_ups = FollowUp.objects.filter(devedor=devedor).order_by('-created_at')

    # Masking CPF or CNPJ
    def mascarar_documento(documento):
        empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
        documento_limpo = re.sub(r'\D', '', documento)
        if len(documento_limpo) == 11:
            return f"{documento_limpo[:3]}.{documento_limpo[3:6]}.xxx.xx"
        elif len(documento_limpo) == 14:
            return f"{documento_limpo[:2]}.{documento_limpo[2:5]}.{documento_limpo[5:6]}xxx-xx"
        return "N/A"

    cpf_cnpj = devedor.cpf or devedor.cnpj
    cpf_cnpj_mascarado = mascarar_documento(cpf_cnpj) if cpf_cnpj else "N/A"

    # Messaging and communications
    # Verifica se o usuário está autenticado
    if request.user.is_authenticated:
        nome_consultor = request.user.get_full_name() or request.user.username
    else:
        nome_consultor = "Usuário Anônimo"
    
    
    mensagem_template = (
        "NOTIFICAÇÃO EXTRAJUDICIAL\n\n"
        "Prezado(a) Cliente {nome},\n"
        "Portador do documento de CPF/CNPJ nº: {cpf_cnpj}. Me chamo {nome_consultor}, falo em nome da SisteAdvAssessoria Assessoria, empresa essa contratada por {nome_credor}.\n\n"
        "Comunicamos a existência de valores não pagos junto à empresa {nome_credor}.\n\n"
        "Solicitamos sua resposta ou comparecimento à loja no prazo de 72 horas após o recebimento dessa Notificação, para tratar de assuntos de sua RESPONSABILIDADE e INTERESSE.\n\n"
        "Entretanto, não sendo e não conhecendo a pessoa acima, pedimos que desconsidere a nossa mensagem.\n\n"
        "Nos colocamos à sua inteira disposição para sanar toda e qualquer dúvida referente a nossa notificação,\n\n"
        "Contato: wa.me://7399840-4699\n"
        "https://api.whatsapp.com/send?phone=5573998404699&text=%20%.\n\n"
        "Estamos à sua disposição.\n"
        "Atenciosamente,\n"
        "Escritório Jurídico\n"
        "Adv Assessoria Jurídica"
    )
    # Definir o nome do consultor fora do método format
    if request.user.is_authenticated:
        nome_consultor = request.user.get_full_name() or request.user.username
    else:
        nome_consultor = "Usuário Anônimo"

    # Construir a mensagem
    
        if request.user.is_authenticated:
            nome_consultor = request.user.get_full_name() or request.user.username
        else:
            nome_consultor = "Usuário Anônimo"
            
    mensagem = mensagem_template.format(
        nome=devedor.nome or "Cliente",
        cpf_cnpj=cpf_cnpj_mascarado,
        nome_consultor=nome_consultor,
        nome_credor=empresa.nome_fantasia if empresa else "NomeCredor"
    )


    # Saving phone changes
    if request.method == 'POST':
        for i in range(1, 11):
            telefone_field = f'telefone{i}'
            telefone_value = request.POST.get(telefone_field)
            if telefone_value:
                setattr(devedor, telefone_field, telefone_value)
        devedor.save()
        #messages.success(request, "Alterações salvas com sucesso!")
        return redirect('detalhes_devedor', titulo_id=titulo.id)

    # Context for template rendering
    context = {
        'devedor': devedor,
        'titulos': titulos,
        'empresa': empresa,
        'agendamentos': agendamentos,
        'follow_ups': follow_ups,
        'telefones': [getattr(devedor, f'telefone{i}', '') for i in range(1, 11)],
        'mensagem_whatsapp': mensagem,
        'total_quitado': total_quitado,
        'total_negociado': total_negociado,
        'total_pendente': total_pendente,
        'forma_pagamento_map': forma_pagamento_map,
        'today': hoje,
        'titulos_entrada': titulos_entrada,
        'titulos_associados': titulos_associados,
    }

    return render(request, 'detalhes_devedor.html', context)
    
@lojista_login_required 
def lista_titulos(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    titulos = Titulo.objects.all()  # Ajuste a lógica conforme necessário
    return render(request, 'lista_titulos.html', {'titulos': titulos})

@lojista_login_required    
def editar_telefones(request, devedor_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    devedor = get_object_or_404(Devedor, id=devedor_id)
    if request.method == "POST":
        devedor.telefone1 = request.POST.get('telefone1')
        devedor.telefone2 = request.POST.get('telefone2')
        devedor.telefone3 = request.POST.get('telefone3')
        devedor.telefone4 = request.POST.get('telefone4')
        devedor.telefone5 = request.POST.get('telefone5')
        devedor.telefone6 = request.POST.get('telefone6')
        devedor.telefone7 = request.POST.get('telefone7')
        devedor.telefone8 = request.POST.get('telefone8')
        devedor.telefone9 = request.POST.get('telefone9')
        devedor.telefone10 = request.POST.get('telefone10')
        devedor.save()
        #messages.success(request, "Telefones atualizados com sucesso!")
        return redirect('detalhes_devedor', titulo_id=devedor.id)
    return redirect('detalhes_devedor', titulo_id=devedor.id)


@lojista_login_required    
def editar_agendamento(request, agendamento_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    devedores = Devedor.objects.all()
    empresas = Empresa.objects.all()  # Fetching all companies
    if request.method == 'POST':
        data = request.POST
        agendamento.devedor = Devedor.objects.get(id=data['devedor_id'])
        agendamento.empresa = Empresa.objects.get(id=data['empresa_id'])
        agendamento.data_abertura = data['data_abertura']
        agendamento.data_retorno = data['data_retorno']
        agendamento.assunto = data['assunto']
        agendamento.operador = data.get('operador', '')
        agendamento.save()
        #messages.success(request, 'Agendamento atualizado com sucesso.')
        return redirect('listar_agendamentos')
    return render(request, 'agendamentos_editar.html', {
        'agendamento': agendamento,
        'devedores': devedores,
        'empresas': empresas  # Use 'empresas' here instead of 'core_empresa'
    })


# Excluir Agendamento
@lojista_login_required    
def excluir_agendamento(request, agendamento_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    agendamento = get_object_or_404(Agendamento, id=agendamento_id)
    agendamento.delete()
    messages.success(request, 'Agendamento excluído com sucesso.')
    return redirect('listar_agendamentos')
 
@lojista_login_required

def listar_follow_ups(request, devedor_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    devedor = get_object_or_404(Devedor, id=devedor_id)
    follow_ups = FollowUp.objects.filter(devedor=devedor).order_by('-created_at')
    return render(request, 'follow_ups_listar.html', {'devedor': devedor, 'follow_ups': follow_ups})

@lojista_login_required
def adicionar_follow_up(request, devedor_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    if request.method == "POST":
        devedor = get_object_or_404(Devedor, id=devedor_id)
        texto = request.POST.get('texto')

        if texto:
            FollowUp.objects.create(devedor=devedor, texto=texto)
            #messages.success(request, "Follow-up adicionado com sucesso.")
        else:
            messages.error(request, "O texto do Follow-up não pode estar vazio.")
        
        # Tentar encontrar um título relacionado
        titulo = Titulo.objects.filter(devedor=devedor).first()
        if titulo:
            return redirect('detalhes_devedor', titulo_id=titulo.id)
        else:
            messages.warning(request, "Nenhum título encontrado para o devedor.")
            return redirect('lista_devedores')  # Substitua por uma view apropriada
    else:
        messages.error(request, "Método inválido.")
        return redirect('lista_devedores')  

def listar_logs(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    # Obtém todos os logs de acesso
    logs = UserAccessLog.objects.all().order_by('-timestamp')
    
    # Cria o paginador
    paginator = Paginator(logs, 30)  # 30 logs por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Passa os logs paginados para o template
    return render(request, 'logs_listar.html', {'page_obj': page_obj}) 

def buscar_devedores(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    if request.method == 'GET':
        termo = request.GET.get('termo', '').strip()
        if termo:
            devedores = Devedor.objects.filter(nome__icontains=termo)[:10]  # Limitar os resultados a 10
            resultados = [
                {
                    "id": devedor.id,
                    "nome": devedor.nome,
                    "empresa_nome": devedor.empresa.nome_fantasia if devedor.empresa else "Não associado"
                }
                for devedor in devedores
            ]
            return JsonResponse(resultados, safe=False)
        return JsonResponse([], safe=False)
        
        
def configurar_permissoes_admin():
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')

    # Obter ou criar o grupo Admin
    admin_group, _ = Group.objects.get_or_create(name='Admin')

    # Associar todas as permissões disponíveis ao grupo Admin
    todas_permissoes = Permission.objects.all()
    admin_group.permissions.set(todas_permissoes)
    admin_group.save()

    print(f"O grupo '{admin_group.name}' agora tem todas as permissões.")
    

def listar_mensagens(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    mensagens = MensagemWhatsapp.objects.all()
    return render(request, 'mensagens_listar.html', {'mensagens': mensagens})

def adicionar_mensagem(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    if request.method == 'POST':
        form = MensagemWhatsappForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_mensagens')
    else:
        form = MensagemWhatsappForm()
    return render(request, 'mensagem_adicionar.html', {'form': form})

def editar_mensagem(request, pk):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    mensagem = get_object_or_404(MensagemWhatsapp, pk=pk)
    if request.method == 'POST':
        form = MensagemWhatsappForm(request.POST, instance=mensagem)
        if form.is_valid():
            form.save()
            return redirect('listar_mensagens')
    else:
        form = MensagemWhatsappForm(instance=mensagem)
    return render(request, 'mensagem_editar.html', {'form': form})

def excluir_mensagem(request, pk):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    mensagem = get_object_or_404(MensagemWhatsapp, pk=pk)
    mensagem.delete()
    return redirect('listar_mensagens')    
    


def tabelas_listar(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    query = request.GET.get('q', '')
    tabelas = TabelaRemuneracao.objects.filter(nome__icontains=query) if query else TabelaRemuneracao.objects.all()
    
    paginator = Paginator(tabelas, 10)  # 10 tabelas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'tabelas_listar.html', {'page_obj': page_obj, 'query': query})



def tabela_adicionar(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    if request.method == 'POST':
        nome = request.POST.get('nome')
        if nome:
            TabelaRemuneracao.objects.create(nome=nome)
            #messages.success(request, "Tabela adicionada com sucesso!")
        return redirect('tabelas_listar')
    return render(request, 'tabela_adicionar.html')


def tabela_editar(request, tabela_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    if request.method == 'POST':
        tabela.nome = request.POST.get('nome')
        tabela.save()
        #messages.success(request, "Tabela editada com sucesso!")
        return redirect('tabelas_listar')
    return render(request, 'tabela_editar.html', {'tabela': tabela})


def tabela_excluir(request, tabela_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    tabela.delete()
    #messages.success(request, "Tabela excluída com sucesso!")
    return redirect('tabelas_listar')


def lista_gerenciar(request, tabela_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)

    if request.method == 'POST':
        # Coleta os dados do formulário
        de_dias = request.POST.get('de_dias')
        ate_dias = request.POST.get('ate_dias')
        percentual_remuneracao = request.POST.get('percentual_remuneracao')

        # Criação de um novo item na tabela
        TabelaRemuneracaoLista.objects.create(
            tabela_remuneracao=tabela,
            de_dias=de_dias,
            ate_dias=ate_dias,
            percentual_remuneracao=percentual_remuneracao
        )

        # Mensagem de sucesso
        #messages.success(request, "Item adicionado à lista!")
    
    # Obtém todos os itens relacionados à tabela
    itens = tabela.listas.all()

    # Renderiza o template com o nome da tabela e os itens
    return render(request, 'lista_gerenciar.html', {'tabela': tabela, 'itens': itens})



def lista_editar(request, tabela_id, item_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    item = get_object_or_404(TabelaRemuneracaoLista, id=item_id, tabela_remuneracao=tabela)

    if request.method == 'POST':
        item.de_dias = request.POST.get('de_dias')
        item.ate_dias = request.POST.get('ate_dias')

        # Substituir vírgula por ponto no percentual_remuneracao
        percentual_remuneracao = request.POST.get('percentual_remuneracao', '').replace(',', '.')
        
        try:
            item.percentual_remuneracao = Decimal(percentual_remuneracao)
            item.save()
            #messages.success(request, "Item atualizado com sucesso!")
            return redirect('lista_gerenciar', tabela_id=tabela.id)
        except Exception as e:
            messages.error(request, f"Erro ao atualizar item: {e}")

    return render(request, 'lista_editar.html', {'tabela': tabela, 'item': item})
    
def lista_adicionar(request, tabela_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    tabela = get_object_or_404(TabelaRemuneracao, id=tabela_id)
    if request.method == 'POST':
        de_dias = request.POST.get('de_dias')
        ate_dias = request.POST.get('ate_dias')
        percentual_remuneracao = request.POST.get('percentual_remuneracao')
        if de_dias and ate_dias and percentual_remuneracao:
            TabelaRemuneracaoLista.objects.create(
                tabela_remuneracao=tabela,
                de_dias=de_dias,
                ate_dias=ate_dias,
                percentual_remuneracao=percentual_remuneracao
            )
            #messages.success(request, "Item adicionado à lista com sucesso!")
        else:
            messages.error(request, "Todos os campos são obrigatórios!")
        return redirect('lista_gerenciar', tabela_id=tabela.id)
    return render(request, 'lista_adicionar.html', {'tabela': tabela})
    
def lista_excluir(request, lista_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    lista = get_object_or_404(TabelaRemuneracaoLista, id=lista_id)
    tabela_id = lista.tabela_remuneracao.id  # Captura o ID da tabela para redirecionar após exclusão
    lista.delete()
    messages.success(request, "Item da lista excluído com sucesso!")
    return redirect('lista_gerenciar', tabela_id=tabela_id)
    


@lojista_login_required
def quitados_listar(request):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão

    # Verifica se o empresa_id_sessao está presente na sessão
    if not empresa_id_sessao:
        messages.error(request, "Sessão expirada ou inválida. Por favor, faça login novamente.")
        return redirect('login')  # Substitua 'login' pelo nome correto da sua URL de login

    # Obter os parâmetros de filtro de data
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()

    # Construir a consulta base com filtro de empresa_id_sessao
    query = """
        SELECT 
            titulo.data_baixa,
            titulo.dataVencimento,
            titulo.valorRecebido,
            devedores.nome,
            devedores.cpf,
            devedores.cnpj,
            core_empresa.nome_fantasia,
            titulo.idTituloRef
        FROM 
            titulo
        INNER JOIN devedores ON titulo.devedor_id = devedores.id
        INNER JOIN core_empresa ON devedores.empresa_id = core_empresa.id
        WHERE 
            titulo.data_baixa is not null and
            devedores.empresa_id = %s
    """

    # Adicionar parâmetros da consulta
    params = [empresa_id_sessao]

    # Filtros adicionais de data
    if data_inicio:
        query += " AND titulo.data_baixa >= %s"
        params.append(data_inicio)
    if data_fim:
        query += " AND titulo.data_baixa <= %s"
        params.append(data_fim)

    # Ordenar os resultados
    query += " ORDER BY titulo.data_baixa DESC;"

    # Executar a consulta
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        rows = cursor.fetchall()

    # Converter os resultados
    quitados = [
        {
            "data_baixa": row[0].strftime('%d/%m/%Y') if row[0] else '',
            "data_vencimento": row[1].strftime('%d/%m/%Y') if row[1] else '',
            "valor_recebido": row[2] or 0.0,  # Tratar None como 0.0
            "nome": row[3],
            "cpf": row[4],
            "cnpj": row[5],
            "empresa": row[6],
            "idTituloRef": row[7],
        }
        for row in rows
    ]

    # Paginação
    paginator = Paginator(quitados, 10)  # 10 itens por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calcular a soma dos valores recebidos
    soma_total = sum(item["valor_recebido"] for item in quitados)

    return render(request, 'quitados.html', {
        'page_obj': page_obj,
        'soma_total': soma_total,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    })
    
@lojista_login_required
def anexar_contrato(request, titulo_id):
    empresa_id_sessao = request.session.get('empresa_id_sessao')  # Pegue o empresa_id da sessão
    empresa_sessao_id = request.session.get('empresa_sessao_id')
    titulo = get_object_or_404(Titulo, pk=titulo_id)
    if request.method == 'POST' and 'contrato' in request.FILES:
        contrato_file = request.FILES['contrato']
        extension = os.path.splitext(contrato_file.name)[1]
        unique_filename = f"{uuid.uuid4()}{extension}"
        titulo.contrato.save(unique_filename, contrato_file)
        titulo.save()
        messages.success(request, "Contrato anexado com sucesso!")
        return redirect('acordos_listar')
    else:
        messages.error(request, "Falha ao anexar o contrato. Tente novamente.")
        return redirect('acordos_listar')



